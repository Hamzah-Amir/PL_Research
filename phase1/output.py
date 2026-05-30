"""
Excel output generation for Phase 1 product shortlists.

Produces a well-formatted workbook with:
  - A header/summary section
  - The ranked product table with conditional score colouring
  - A filter-settings reference sheet
  - A scoring legend
"""

import urllib.request
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from phase1.filters import FilterConfig

# Embedded product-image thumbnail size (pixels). Row height / column width
# in the Shortlist sheet are sized to fit this.
_IMG_PX = 80

# ──────────────────────────────────────────────────────────────────────────────
# Colour palette
# ──────────────────────────────────────────────────────────────────────────────
C = {
    "hdr_bg":      "C00000",   # Dark red
    "hdr_fg":      "FFFFFF",
    "sub_bg":      "1F3864",   # Dark navy
    "sub_fg":      "FFFFFF",
    "row_odd":     "EFF3FF",   # Very light blue
    "row_even":    "FFFFFF",
    "score_hi":    "70AD47",   # Green  ≥ 70
    "score_mid":   "FFC000",   # Amber  50–69
    "score_lo":    "FF7C80",   # Red    < 50
    "flag_cell":   "FFE699",   # Yellow for seasonal/risk highlight
    "border":      "BFBFBF",
}

# ──────────────────────────────────────────────────────────────────────────────
# Column definitions: (canonical_key, display_label, width)
# ──────────────────────────────────────────────────────────────────────────────
COLUMNS: List[Tuple[str, str, int]] = [
    ("asin",               "ASIN",                 13),
    ("image_url",          "Image",                 13),
    ("title",              "Product Name",          48),
    ("brand",              "Brand",                 18),
    ("category",           "Category",              22),
    ("price",              "Price (£)",             10),
    ("monthly_revenue",    "Monthly Revenue (£)",   18),
    ("monthly_sales",      "Monthly Sales",         14),
    ("bsr",                "BSR",                   10),
    ("rating",             "Rating",                 8),
    ("reviews",            "Reviews",               10),
    ("listing_age_months", "Listing Age (mo)",      14),
    ("fulfillment",        "Fulfillment",           12),
    ("composite_score",    "Score",                  8),
    ("risk_flags",         "Notes / Risk Flags",    42),
]


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold: bool = False, size: int = 9, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _border_thin() -> Border:
    side = Side(style="thin", color=C["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def _safe_float(value) -> Optional[float]:
    if value is None:
        return None
    try:
        f = float(value)
        return None if np.isnan(f) else f
    except (ValueError, TypeError):
        return None


def _safe_int(value) -> Optional[int]:
    f = _safe_float(value)
    return None if f is None else int(round(f))


def _score_color(score: Optional[float]) -> str:
    if score is None:
        return C["row_odd"]
    if score >= 70:
        return C["score_hi"]
    if score >= 50:
        return C["score_mid"]
    return C["score_lo"]


def _make_thumbnail(url: str, cache: Dict[str, Optional[BytesIO]]) -> Optional[BytesIO]:
    """
    Download an image URL and return a PNG thumbnail as an in-memory buffer,
    ready to hand to openpyxl. Returns None on any failure (network, bad URL,
    unreadable image) so the export never crashes. Results are cached per URL.
    """
    if not url or not str(url).lower().startswith(("http://", "https://")):
        return None
    if url in cache:
        # Return a fresh copy of the cached bytes (openpyxl consumes the buffer)
        cached = cache[url]
        return BytesIO(cached.getvalue()) if cached is not None else None

    buf: Optional[BytesIO] = None
    try:
        from PIL import Image as PILImage  # local import — optional dependency

        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read()

        im = PILImage.open(BytesIO(raw)).convert("RGB")
        im.thumbnail((_IMG_PX, _IMG_PX))
        buf = BytesIO()
        im.save(buf, format="PNG")
        buf.seek(0)
    except Exception:
        buf = None

    cache[url] = buf
    return BytesIO(buf.getvalue()) if buf is not None else None


def _embed_image(
    ws,
    anchor: str,
    url: str,
    cache: Dict[str, Optional[BytesIO]],
) -> bool:
    """Embed a product thumbnail at *anchor* (e.g. 'B4'). Returns True on success."""
    thumb = _make_thumbnail(url, cache)
    if thumb is None:
        return False
    try:
        img = XLImage(thumb)
        img.width = _IMG_PX
        img.height = _IMG_PX
        ws.add_image(img, anchor)
        return True
    except Exception:
        return False


# ──────────────────────────────────────────────────────────────────────────────
# Sheet builders
# ──────────────────────────────────────────────────────────────────────────────

def _write_products_sheet(
    wb: openpyxl.Workbook,
    batch: pd.DataFrame,
    batch_num: int,
    category: str,
    budget: float,
    max_revenue: float,
    variant_label: str = "",
) -> None:
    ws = wb.active
    title_suffix = f" — {variant_label}" if variant_label else ""
    ws.title = (f"B{batch_num}{title_suffix}")[:31]

    columns = COLUMNS
    num_cols = len(columns)
    last_col_letter = get_column_letter(num_cols)

    # ── Row 1: main title ────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = (
        f"PL Research Tool  |  Phase 1: Product Hunting  |  "
        f"Batch {batch_num}{title_suffix}"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C["hdr_fg"])
    ws["A1"].fill = _fill(C["hdr_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: metadata strip ────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = (
        f"Category: {category}   |   Budget: £{budget:,.0f}   |   "
        f"Revenue Cap: £{max_revenue:,.0f}/mo   |   "
        f"Current month: {datetime.now().strftime('%B %Y')}   |   "
        f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, color=C["sub_fg"])
    ws["A2"].fill = _fill(C["sub_bg"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Row 3: column headers ────────────────────────────────────────────────
    for col_idx, (_, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=C["sub_fg"])
        cell.fill = _fill(C["sub_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 30

    # ── Rows 4+: product data ────────────────────────────────────────────────
    img_cache: Dict[str, Optional[BytesIO]] = {}
    for row_offset, (_, prod) in enumerate(batch.iterrows()):
        row_num = 4 + row_offset
        base_fill = C["row_odd"] if row_offset % 2 == 0 else C["row_even"]

        for col_idx, (key, _, _) in enumerate(columns, start=1):
            cell = ws.cell(row=row_num, column=col_idx)
            raw = prod.get(key)

            if key == "image_url":
                # Style the cell, then embed the downloaded thumbnail on top.
                cell.fill = _fill(base_fill)
                cell.border = _border_thin()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                anchor = f"{get_column_letter(col_idx)}{row_num}"
                if not _embed_image(ws, anchor, str(raw) if pd.notna(raw) else "", img_cache):
                    cell.value = "no image"
                    cell.font = _font(size=8, color="999999")
                continue

            if key == "price":
                cell.value = _safe_float(raw)
                cell.number_format = '£#,##0.00'
            elif key == "monthly_revenue":
                cell.value = _safe_float(raw)
                cell.number_format = '£#,##0'
            elif key in ("monthly_sales", "reviews", "bsr", "variations"):
                cell.value = _safe_int(raw)
                cell.number_format = '#,##0'
            elif key == "rating":
                cell.value = _safe_float(raw)
                cell.number_format = '0.0'
            elif key == "listing_age_months":
                cell.value = _safe_int(raw)
            elif key == "composite_score":
                score = _safe_float(raw)
                cell.value = score
                cell.number_format = '0.0'
                cell.fill = _fill(_score_color(score))
                cell.font = Font(name="Calibri", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _border_thin()
                continue
            else:
                cell.value = str(raw) if (raw is not None and pd.notna(raw)) else ""

            # Risk-flag column: highlight if non-empty
            if key == "risk_flags" and cell.value and cell.value.strip() not in ("", "None"):
                cell.fill = _fill(C["flag_cell"])
            else:
                cell.fill = _fill(base_fill)

            cell.font = _font()
            cell.alignment = Alignment(
                horizontal="left" if key in ("title", "risk_flags") else "center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = _border_thin()

        # Tall enough to fit the embedded image thumbnail (px → points ≈ ×0.75)
        ws.row_dimensions[row_num].height = max(38, _IMG_PX * 0.75 + 2)

    # ── Legend rows ──────────────────────────────────────────────────────────
    legend_row = 4 + len(batch) + 1
    ws.merge_cells(f"A{legend_row}:{last_col_letter}{legend_row}")
    ws[f"A{legend_row}"] = "SCORE LEGEND"
    ws[f"A{legend_row}"].font = Font(name="Calibri", bold=True, size=9, color=C["sub_fg"])
    ws[f"A{legend_row}"].fill = _fill(C["sub_bg"])
    ws[f"A{legend_row}"].alignment = Alignment(horizontal="left")

    legends = [
        (C["score_hi"],  "Score ≥ 70  — Strong candidate: high revenue, good rating, proven reviews"),
        (C["score_mid"], "Score 50–69 — Moderate candidate: solid numbers, minor concerns"),
        (C["score_lo"],  "Score < 50  — Weaker candidate: review below the given ranking threshold"),
        (C["flag_cell"], "Yellow cell in 'Notes' — one or more risk flags detected (review before selecting)"),
    ]
    for i, (color, text) in enumerate(legends):
        r = legend_row + 1 + i
        ws.cell(row=r, column=1).fill = _fill(color)
        ws.merge_cells(f"B{r}:{last_col_letter}{r}")
        ws[f"B{r}"] = text
        ws[f"B{r}"].font = _font(size=9)
        ws.row_dimensions[r].height = 16

    # Freeze header rows
    ws.freeze_panes = "A4"


def _write_settings_sheet(
    wb: openpyxl.Workbook,
    config: FilterConfig,
    budget: float,
) -> None:
    """Add a 'Filter Settings' sheet for reference."""
    ws = wb.create_sheet("Filter Settings")

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:B1")
    ws["A1"] = "Phase 1 — Filter Settings Applied"
    ws["A1"].font = Font(name="Calibri", bold=True, size=12, color=C["hdr_fg"])
    ws["A1"].fill = _fill(C["hdr_bg"])
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    def _price_label(cfg) -> str:
        return f"≥ £{cfg.f5_min_price:,.2f} (no upper limit)"

    seasonal_status = (
        "Included (summer/unknown kept & flagged; Christmas/Q4 removed)"
        if config.want_seasonal
        else "Excluded (all seasonal removed)"
    )

    if config.min_revenue > 0:
        revenue_value = (
            f"£{config.min_revenue:,.0f} – £{config.max_revenue:,.0f}/month  "
            f"(source: ASIN-level revenue; cap = budget × 1.52)"
        )
    else:
        revenue_value = (
            f"≤ £{config.max_revenue:,.0f}/month  "
            f"(source: ASIN-level revenue; cap = budget × 1.52, no minimum)"
        )

    rows = [
        ("Budget", f"£{budget:,.0f}"),
        ("Revenue filter", revenue_value),
        ("─" * 30, "─" * 40),
        ("F1  Amazon-as-seller", "Removed" if config.f1_remove_amazon else "Kept"),
        ("F2  Seasonal products", seasonal_status),
        ("F3  Min listing age", f"{config.f3_min_listing_age_months} months"),
        ("F4  Min rating", f"> {config.f4_min_rating} stars"),
        ("F5  Min unit price", _price_label(config)),
        (
            "F6  90-day sales trend",
            f"Remove if decline > {abs(config.f6_max_decline_pct):.0f}% over 90 days"
            if config.f6_apply else "Not applied",
        ),
        ("F7  Max reviews", f"≤ {config.f7_max_reviews:,}"),
        (
            "F8  Regulated exclusions",
            ", ".join(config.f8_excluded_categories)
            if config.f8_excluded_categories else "None",
        ),
        ("F9  Compatibility products", "Removed" if config.f9_remove_compatibility else "Kept"),
    ]

    for r_idx, (label, value) in enumerate(rows, start=2):
        a = ws.cell(row=r_idx, column=1, value=label)
        b = ws.cell(row=r_idx, column=2, value=value)
        for cell in (a, b):
            cell.font = _font(size=10)
            cell.alignment = Alignment(vertical="center")
        if label.startswith("─"):
            a.fill = _fill("E0E0E0")
            b.fill = _fill("E0E0E0")
        ws.row_dimensions[r_idx].height = 16


# ──────────────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────────────

def export_to_excel(
    batch: pd.DataFrame,
    batch_num: int,
    category: str,
    budget: float,
    max_revenue: float,
    filter_config: FilterConfig,
    variant_label: str = "",
    filename_tag: str = "",
) -> str:
    """
    Write the product batch to a formatted Excel file.

    *variant_label* (e.g. "With Variants") is shown in the sheet header/title;
    *filename_tag* (e.g. "WithVariants") is embedded in the file name.

    Returns the path to the saved file as a string.
    """
    output_dir = Path("output")
    output_dir.mkdir(exist_ok=True)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    tag = f"_{filename_tag}" if filename_tag else ""
    filename = f"Phase1_Batch{batch_num}{tag}_{ts}.xlsx"
    filepath = output_dir / filename

    wb = openpyxl.Workbook()

    _write_products_sheet(
        wb, batch, batch_num, category, budget, max_revenue, variant_label,
    )
    _write_settings_sheet(wb, filter_config, budget)

    wb.save(filepath)
    return str(filepath)


def print_product_table(batch: pd.DataFrame, batch_num: int, variant_label: str = "") -> None:
    """Print a compact product summary table to the terminal."""
    label = f"  [{variant_label}]" if variant_label else ""
    print(f"\n  ── Top {len(batch)} Products — Batch {batch_num}{label} {'─'*25}")
    header = (
        f"  {'#':>2}  {'ASIN':<12}  {'Brand':<16}  "
        f"{'Price':>7}  {'Sales':>6}  {'Revenue':>10}  "
        f"{'Rating':>6}  {'Reviews':>7}  {'Score':>5}  Risk"
    )
    print(header)
    print(f"  {'─'*120}")

    for rank, (_, row) in enumerate(batch.iterrows(), start=1):
        asin = str(row.get("asin", ""))[:12]
        brand = str(row.get("brand", ""))[:16]
        price = row.get("price")
        sales = row.get("monthly_sales")
        revenue = row.get("monthly_revenue")
        rating = row.get("rating")
        reviews = row.get("reviews")
        score = row.get("composite_score")
        flags = str(row.get("risk_flags", "None"))[:35]

        print(
            f"  {rank:>2}.  {asin:<12}  {brand:<16}  "
            f"{('£'+f'{price:,.0f}') if price else 'N/A':>7}  "
            f"{int(sales) if sales else 'N/A':>6}  "
            f"{('£'+f'{revenue:,.0f}') if revenue else 'N/A':>10}  "
            f"{f'{rating:.1f}' if rating else 'N/A':>6}  "
            f"{int(reviews) if reviews else 'N/A':>7}  "
            f"{f'{score:.1f}' if score else 'N/A':>5}  "
            f"{flags}"
        )
