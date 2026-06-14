"""
Phase 3 deliverable writer — fills the Competitor Analysis workbook
(Competitor Analysis + Pricing Analysis + Sponsored Products tabs).

Design choice: the grading rows are located by their LABEL TEXT in column E
within each competitor block and section — never by hard-coded row numbers — so
the writer survives the user editing the template (e.g. adding the 4 new
"MARKET MOMENTUM" rows, renaming "AMS Ads" -> "Sponsored Brands"). For each
matched element we write the Info cell (col F) and the Score cell (col J); the
template's `Total Score = SUM(J..)` picks the scores up. Unknown scores
(Rule 0) write "Unknown — need user confirmation" in Info and leave Score blank.

The template lives in test_file/; output is always written to a new timestamped
file in output/ (the template is never overwritten).
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

# Column letters in the Competitor Analysis tab.
COL_GRADING, COL_INFO, COL_SCORE, COL_MAX = "E", "F", "J", "K"
COL_META_LABEL, COL_META_VALUE = "B", "C"

# Map a normalised column-E label -> scorer element key. Marketing-section
# duplicates (images/title/bullets) are disambiguated by section state.
_LABEL_KEYS = {
    "price": "price",
    "sponsored products": "sponsored_products",
    "ams ads": "sponsored_brands",
    "sponsored brands": "sponsored_brands",
    "sponsored brand": "sponsored_brands",
    "# of product reviews": "reviews",
    "avg. product rating": "rating",
    "age of product": "age",
    "bestseller badge": "bestseller",
    "amazon or 3p seller": "seller",
    "unique product design": "unique_design",
    "fba or fbm": "fba_fbm",
    "pricing strategy": "pricing_strategy",
    # new MARKET MOMENTUM rows:
    "review velocity": "review_velocity",
    "product variations": "variations",
    "sales / revenue strength": "sales_strength",
    "sales/revenue strength": "sales_strength",
    "enhanced content (a+/video)": "enhanced_content",
    "enhanced content": "enhanced_content",
}
# Section-sensitive labels: (product-page key, marketing key)
_SECTION_LABEL_KEYS = {
    "product images": ("product_images", "mkt_images"),
    "product title": ("product_title", "mkt_title"),
    "product description/bullet points": ("bullets_description", "mkt_bullets"),
}

_META_KEYS = {
    "brand name": "brand",
    "product page url": "url",
    "seller name": "seller_name",
    "units sales/month": "monthly_sales",
    "parent category": "category",
}


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower()).rstrip(":")


def _find_block_starts(ws) -> List[int]:
    """Rows where each competitor block begins (col B '..Competitor')."""
    starts = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b and re.search(r"\bcompetitor\b", str(b), re.I) and re.match(r"\s*\d", str(b)):
            starts.append(r)
    return starts


def _info_fallback(key: str, comp: Dict) -> Optional[str]:
    """Info-cell content for judgement elements whose scorer left no info text:
    the actual title for the Product Title rows, the media facts for the
    Product Images rows."""
    if key in ("product_title", "mkt_title"):
        return comp.get("title")
    if key in ("product_images", "mkt_images"):
        n = comp.get("images_count")
        if n is None:
            return None
        return f"{int(n)} listing image(s)" + (", video present" if comp.get("has_video") else "")
    if key in ("bullets_description", "mkt_bullets"):
        bullets = [str(b).strip() for b in (comp.get("bullets") or []) if str(b or "").strip()]
        if bullets:
            return "\n".join(f"- {b}" for b in bullets[:5])[:600]
        desc = (comp.get("description") or "").strip()
        return desc[:600] or None
    return None


def _embed_product_image(ws, row: int, comp: Dict) -> bool:
    """Embed the main product image in the Info cell of a Product Images
    grading row, scaled to the row height. Returns True on success."""
    urls = list(comp.get("image_urls") or [])
    if not urls and comp.get("image_url"):
        urls = [comp["image_url"]]
    if not urls:
        return False
    try:
        import io
        import requests
        from openpyxl.drawing.image import Image as XLImage
        resp = requests.get(urls[0], timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        img = XLImage(io.BytesIO(resp.content))
        row_pt = ws.row_dimensions[row].height or 60
        target = max(40.0, row_pt * 4 / 3 - 6)  # row height (pt) -> px, small margin
        scale = min(target / img.height, 1.0)
        img.width, img.height = int(img.width * scale), int(img.height * scale)
        ws.add_image(img, f"{COL_INFO}{row}")
        return True
    except Exception:  # noqa: BLE001 — caller falls back to text info
        return False


def _fill_product_box(ws, start: int, end: int, comp: Dict) -> None:
    """The tall empty merged box under the 'N Competitor' label (B11:C17 etc.)
    gets the PRODUCT TITLE (user decision — no image here)."""
    if not comp.get("title"):
        return
    for mr in sorted(ws.merged_cells.ranges, key=lambda m: m.min_row):
        if (mr.min_col == 2 and mr.max_col >= 3 and start < mr.min_row
                and mr.max_row < end and (mr.max_row - mr.min_row) >= 3):
            from openpyxl.styles import Alignment
            anchor = ws.cell(mr.min_row, mr.min_col)
            anchor.value = comp["title"]
            anchor.alignment = Alignment(wrap_text=True, vertical="center",
                                         horizontal="center")
            return


def _fill_block(ws, start: int, end: int, comp: Dict, scored: Dict, review) -> List[str]:
    """Fill one competitor block in [start, end). Returns notes about misses."""
    misses: List[str] = []
    elements = scored["elements"]
    in_marketing = False
    seen_keys = set()

    _fill_product_box(ws, start, end, comp)

    for r in range(start, end):
        # Section transition
        gtext = " ".join(
            _norm(ws.cell(row=r, column=c).value) for c in (5, 6, 7)
        )
        if "overall marketing strategy" in gtext:
            in_marketing = True

        # Metadata (col B label -> col C value)
        blabel = _norm(ws.cell(row=r, column=2).value)
        if blabel in _META_KEYS:
            key = _META_KEYS[blabel]
            if key == "seller_name":
                val = comp.get("seller") or comp.get("buy_box")
            elif key == "category":
                val = comp.get("subcategory") or comp.get("category")
            else:
                val = comp.get(key)
            if val is not None:
                ws[f"{COL_META_VALUE}{r}"] = val

        # Grading element (col E label)
        elabel = _norm(ws.cell(row=r, column=5).value)
        key = None
        if elabel in _SECTION_LABEL_KEYS:
            key = _SECTION_LABEL_KEYS[elabel][1 if in_marketing else 0]
        elif elabel in _LABEL_KEYS:
            key = _LABEL_KEYS[elabel]
        if key and key in elements and key not in seen_keys:
            seen_keys.add(key)
            el = elements[key]
            if el["score"] is None:
                ws[f"{COL_INFO}{r}"] = el.get("info") or "Unknown — need user confirmation"
                ws[f"{COL_SCORE}{r}"] = None
            else:
                # The Price Info cell feeds numeric formulas in Pricing Analysis
                # (Average Retail Price -> FBA payout), so write the raw number.
                if key == "price" and comp.get("price") is not None:
                    info_cell = ws[f"{COL_INFO}{r}"]
                    info_cell.value = round(float(comp["price"]), 2)
                    info_cell.number_format = "£#,##0.00"
                elif key in ("product_images", "mkt_images") and _embed_product_image(ws, r, comp):
                    pass  # the image itself is the Info content
                elif el.get("info"):
                    ws[f"{COL_INFO}{r}"] = el["info"]
                else:
                    fb = _info_fallback(key, comp)
                    if fb:
                        ws[f"{COL_INFO}{r}"] = fb
                ws[f"{COL_SCORE}{r}"] = el["score"]

    # Weaknesses / solutions / strengths (label headers -> the "1./2./3." rows)
    _fill_review_section(ws, start, end, review)

    # Report any rubric element we never found a row for (e.g. new rows not added)
    missing_keys = [k for k in elements if k not in seen_keys]
    if missing_keys:
        misses.append(f"block@{start}: no row for {', '.join(missing_keys)}")
    return misses


def _is_review_header(e: str) -> Optional[str]:
    """Return 'weak' / 'sol' / 'strong' if E-label *e* is a review-section header."""
    if "solutions to top 3" in e:
        return "sol"  # must be checked first: contains "top 3 weakness" too
    if "top 3 weakness" in e:
        return "weak"
    if "top 3 strength" in e:
        return "strong"
    return None


def _fill_review_section(ws, start: int, end: int, review) -> None:
    """Write the 3 weaknesses / solutions / strengths under their headers.

    Each header (e.g. 'TOP 3 WEAKNESSES') is followed by '1.'/'2.'/'3.' rows in
    column E; we append the text to those. With no review data the first item
    carries the Unknown note (Rule 0).
    """
    lists = {
        "weak": list(getattr(review, "weaknesses", []) or []),
        "sol": list(getattr(review, "solutions", []) or []),
        "strong": list(getattr(review, "strengths", []) or []),
    }
    note = getattr(review, "note", "") or ""

    for r in range(start, end):
        kind = _is_review_header(_norm(ws.cell(row=r, column=5).value))
        if not kind:
            continue
        items = lists[kind]
        filled = 0
        rr = r + 1
        while rr < end and filled < 3:
            cell_e = str(ws.cell(row=rr, column=5).value or "").strip()
            if _is_review_header(_norm(cell_e)):
                break
            m = re.match(r"^([123])\.", cell_e)
            if m:
                idx = int(m.group(1)) - 1
                cell = ws.cell(row=rr, column=5)
                if idx < len(items):
                    cell.value = f"{idx + 1}. {items[idx]}"
                elif not items and idx == 0 and note:
                    cell.value = f"1. {note}"
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                filled += 1
            rr += 1


def _fill_pricing_analysis(wb, comps: List[Dict], target_title: Optional[str] = None) -> None:
    """Clear Rule-0-violating placeholder costs and write the data we have.

    The competitor rows (13/14/15) pull Parent Category / URL / Brand / Price /
    Sales from the Competitor Analysis tab via existing formulas — those populate
    automatically. Here we (a) blank COGS/shipping/HTS placeholders (Rule 0 —
    blank until sourcing) and (b) write BSR, dimensions and weight from our data.
    """
    if "Pricing Analysis" not in wb.sheetnames:
        return
    ws = wb["Pricing Analysis"]

    # The TARGET product's title (from Phase 1) replaces the 'Electric Balloon
    # Pump' placeholder in the merged D11:F11 header cell.
    if target_title:
        target_cell = "D11"
        for row in ws.iter_rows(min_row=5, max_row=15):
            for c in row:
                if c.value and "balloon" in str(c.value).lower():
                    target_cell = c.coordinate
                    break
        ws[target_cell] = target_title

    # Repair the Top-3 row's template refs (#REF! in the original template) so
    # category/URL/brand/price/sales flow from the 3rd competitor block too.
    _BLOCK3_REFS = {
        "D15": "=+'Competitor Analysis'!C115",  # Parent Category
        "E15": "=+'Competitor Analysis'!C112",  # Product URL
        "F15": "=+'Competitor Analysis'!C111",  # Brand / Product Name
        "I15": "='Competitor Analysis'!F106",   # Avg Retail Price (price Info cell)
        "L15": "='Competitor Analysis'!C114",   # Projected Monthly Sales
    }
    for cell, formula in _BLOCK3_REFS.items():
        if isinstance(ws[cell].value, str) and "#REF" in ws[cell].value:
            ws[cell] = formula

    rows = [13, 14, 15]  # Top 1 / 2 / 3 competitor rows
    # Columns: G=Avg BSR, H=COGS, I=Avg Retail Price, J=FBA payout (sea),
    #          L=Projected sales, M/N=freight per unit, P=Shipping,
    #          X/Y=HTS, AB/AC/AD=Size cm, AG=Weight kg.
    for i, r in enumerate(rows):
        comp = comps[i] if i < len(comps) else None
        # Clear Rule-0 cost placeholders.
        for col in ("H", "P", "X", "Y"):
            ws[f"{col}{r}"] = None
        if comp is None:
            for col in ("G", "M", "N"):
                ws[f"{col}{r}"] = None
            continue
        # Average Parent BSR Rank = Keepa 90-day average (current BSR fallback).
        bsr = comp.get("bsr_90d")
        if not isinstance(bsr, (int, float)):
            bsr = comp.get("bsr")
        ws[f"G{r}"] = int(bsr) if isinstance(bsr, (int, float)) else None
        # FBA payout (header formula), with THIS competitor's actual Keepa fees
        # instead of the example product's hardcoded 15.3% / £2.87. The +0.5
        # supplier->warehouse shipping placeholder is kept from the template.
        pct, fee = comp.get("referral_pct"), comp.get("fba_pick_pack_fee")
        if isinstance(pct, (int, float)) and isinstance(fee, (int, float)):
            ws[f"J{r}"] = (f"=I{r}-((I{r}*{round(pct / 100.0, 4)})"
                           f"+(I{r}/6)+({round(fee, 2)})+(0.5))")
        dims = _parse_dims_cm(comp.get("dimensions"))
        if dims:
            ws[f"AB{r}"], ws[f"AC{r}"], ws[f"AD{r}"] = dims
        wt = _parse_weight_kg(comp.get("weight"), comp.get("package_weight_g"))
        if wt is not None:
            ws[f"AG{r}"] = wt
            # Header formula per unit: (sales x weight x 1.1 x rate)/sales -> the
            # sales terms cancel, leaving weight x 1.1 x $3 (sea) / $6 (air).
            ws[f"M{r}"] = f"=AG{r}*1.1*3"
            ws[f"N{r}"] = f"=AG{r}*1.1*6"
        else:
            ws[f"M{r}"], ws[f"N{r}"] = None, None  # clear the template's =26/#REF!


def _parse_dims_cm(value) -> Optional[List[float]]:
    if not value:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", str(value))
    if len(nums) >= 3:
        return [float(nums[0]), float(nums[1]), float(nums[2])]
    return None


def _parse_weight_kg(weight_str, package_weight_g) -> Optional[float]:
    if weight_str:
        s = str(weight_str).lower()
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        if m:
            val = float(m.group(1))
            return val if "kg" in s else round(val / 1000.0, 3) if "g" in s else val
    if isinstance(package_weight_g, (int, float)) and package_weight_g > 0:
        return round(package_weight_g / 1000.0, 3)
    return None


def _fill_sponsored_products(wb, keywords: List[str]) -> None:
    """Write up to 6 launch keywords into the Sponsored Products tab, for each
    match type (Exact rows 7-12, Phrase 14-19, Broad 21-26). Bids stay blank."""
    if "Sponsored Products" not in wb.sheetnames:
        return
    ws = wb["Sponsored Products"]
    blocks = {"exact": 7, "phrase": 14, "broad": 21}
    kws = (keywords or [])[:6]
    for _match, start in blocks.items():
        for i, kw in enumerate(kws):
            ws.cell(row=start + i, column=4, value=kw)  # col D = Keyword


def write_competitor_workbook(
    template_path: str,
    top: List[Dict],
    scored_by_asin: Dict[str, Dict],
    reviews_by_asin: Dict[str, object],
    keywords: List[str],
    target_title: Optional[str] = None,
) -> str:
    """
    Fill the Competitor Analysis template for the top-N competitors and save a
    new timestamped workbook in output/. *top* is an ordered list of merged
    competitor record dicts (block 1 = top[0], etc.). Returns the saved path.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Competitor Analysis"]
    starts = _find_block_starts(ws)
    if not starts:
        raise ValueError("Could not locate competitor blocks in the template.")
    starts.append(ws.max_row + 1)  # sentinel end

    all_misses: List[str] = []
    for i, comp in enumerate(top):
        if i >= len(starts) - 1:
            break
        start, end = starts[i], starts[i + 1]
        scored = scored_by_asin.get(comp["asin"])
        review = reviews_by_asin.get(comp["asin"])
        if scored:
            all_misses += _fill_block(ws, start, end, comp, scored, review)

    _fill_pricing_analysis(wb, top, target_title)
    _fill_sponsored_products(wb, keywords)

    out_dir = Path("output")
    out_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"Phase3_CompetitorAnalysis_{ts}.xlsx"
    wb.save(out_path)

    if all_misses:
        print("  NOTE (template rows not found — add them to fill):")
        for m in all_misses:
            print(f"    - {m}")
    return str(out_path)
