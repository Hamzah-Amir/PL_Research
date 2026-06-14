"""
Excel output for the Phase 1 shortlist (DB-backed model).

Renders one ranked batch of products (e.g. the top 10 of the opportunity pool)
to a formatted workbook with embedded product thumbnails, an opportunity-score
colour scale, and a small "Query" sheet recording the budget / filters used.
Styling helpers are shared with the previous Phase 1 build.
"""

import urllib.request
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_IMG_PX = 80

C = {
    "hdr_bg": "C00000", "hdr_fg": "FFFFFF",
    "sub_bg": "1F3864", "sub_fg": "FFFFFF",
    "row_odd": "EFF3FF", "row_even": "FFFFFF",
    "score_hi": "70AD47", "score_mid": "FFC000", "score_lo": "FF7C80",
    "flag_cell": "FFE699", "border": "BFBFBF",
}

# (key, label, width)
COLUMNS: List[Tuple[str, str, int]] = [
    ("rank", "#", 5),
    ("asin", "ASIN", 13),
    ("image_url", "Image", 12),
    ("title", "Product Name", 46),
    ("brand", "Brand", 16),
    ("subcategory", "Sub-Category", 20),
    ("price", "Price (£)", 10),
    ("estimated_revenue", "ASIN Revenue (£/mo)", 17),
    ("estimated_sales", "ASIN Sales/mo", 13),
    ("bsr", "BSR", 9),
    ("rating", "Rating", 7),
    ("review_count", "Reviews", 9),
    ("variation_count", "Variants", 9),
    ("weight_kg", "Weight kg", 9),
    ("yoy_growth", "YoY %", 8),
    ("sales_trend_90d", "Trend90 %", 9),
    ("revenue_proximity", "Rev Fit", 8),
    ("opportunity_final", "Opportunity", 11),
    ("llm_adj", "LLM Adj", 8),
    ("llm_reason", "LLM Reason", 30),
    ("is_seasonal", "Seasonal", 9),
]


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, size=9, color="000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _border_thin() -> Border:
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)


def _f(value) -> Optional[float]:
    if value is None:
        return None
    try:
        x = float(value)
        return None if np.isnan(x) else x
    except (ValueError, TypeError):
        return None


def _i(value) -> Optional[int]:
    x = _f(value)
    return None if x is None else int(round(x))


def _score_color(score: Optional[float]) -> str:
    # opportunity_final is out of ~115 (base 85 + revenue proximity 30)
    if score is None:
        return C["row_odd"]
    if score >= 80:
        return C["score_hi"]
    if score >= 55:
        return C["score_mid"]
    return C["score_lo"]


def _make_thumbnail(url: str, cache: Dict[str, Optional[BytesIO]]) -> Optional[BytesIO]:
    if not url or not str(url).lower().startswith(("http://", "https://")):
        return None
    if url in cache:
        c = cache[url]
        return BytesIO(c.getvalue()) if c is not None else None
    buf = None
    try:
        from PIL import Image as PILImage
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read()
        im = PILImage.open(BytesIO(raw)).convert("RGB")
        im.thumbnail((_IMG_PX, _IMG_PX))
        buf = BytesIO()
        im.save(buf, format="PNG")
        buf.seek(0)
    except Exception:
        buf = None
    cache[url] = buf
    return BytesIO(buf.getvalue()) if buf is not None else None


def _embed_image(ws, anchor, url, cache) -> bool:
    thumb = _make_thumbnail(url, cache)
    if thumb is None:
        return False
    try:
        img = XLImage(thumb)
        img.width = _IMG_PX
        img.height = _IMG_PX
        ws.add_image(img, anchor)
        return True
    except Exception:
        return False


def _write_products_sheet(wb, products: List[Dict], batch_num: int, meta: str) -> None:
    ws = wb.active
    ws.title = f"Shortlist B{batch_num}"[:31]
    ncols = len(COLUMNS)
    last = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last}1")
    ws["A1"] = f"PL Research Tool  |  Phase 1: Product Shortlist  |  Batch {batch_num}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=C["hdr_fg"])
    ws["A1"].fill = _fill(C["hdr_bg"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    ws.merge_cells(f"A2:{last}2")
    ws["A2"] = meta
    ws["A2"].font = Font(name="Calibri", size=10, color=C["sub_fg"])
    ws["A2"].fill = _fill(C["sub_bg"])
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for ci, (_, label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=ci, value=label)
        cell.font = Font(name="Calibri", bold=True, size=9, color=C["sub_fg"])
        cell.fill = _fill(C["sub_bg"])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border_thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 30

    img_cache: Dict[str, Optional[BytesIO]] = {}
    for ro, prod in enumerate(products):
        row = 4 + ro
        base = C["row_odd"] if ro % 2 == 0 else C["row_even"]
        for ci, (key, _, _) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row, column=ci)
            if key == "rank":
                cell.value = ro + 1
                cell.fill = _fill(base); cell.border = _border_thin()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = _font(bold=True)
                continue
            if key == "image_url":
                cell.fill = _fill(base); cell.border = _border_thin()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if not _embed_image(ws, f"{get_column_letter(ci)}{row}", str(prod.get("image_url") or ""), img_cache):
                    cell.value = "no image"; cell.font = _font(size=8, color="999999")
                continue
            raw = prod.get(key)
            if key == "price":
                cell.value = _f(raw); cell.number_format = '£#,##0.00'
            elif key == "estimated_revenue":
                cell.value = _f(raw); cell.number_format = '£#,##0'
            elif key in ("estimated_sales", "review_count", "bsr", "variation_count"):
                cell.value = _i(raw); cell.number_format = '#,##0'
            elif key == "rating":
                cell.value = _f(raw); cell.number_format = '0.0'
            elif key == "weight_kg":
                cell.value = _f(raw); cell.number_format = '0.00'
            elif key in ("yoy_growth", "sales_trend_90d", "revenue_proximity"):
                cell.value = _f(raw); cell.number_format = '0.0'
            elif key == "opportunity_final":
                sc = _f(raw)
                cell.value = sc; cell.number_format = '0.0'
                cell.fill = _fill(_score_color(sc))
                cell.font = Font(name="Calibri", size=9, bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = _border_thin()
                continue
            elif key == "is_seasonal":
                cell.value = "Yes" if raw else "No"
            else:
                cell.value = str(raw) if raw is not None and not (isinstance(raw, float) and np.isnan(raw)) else ""

            cell.fill = _fill(C["flag_cell"] if (key == "is_seasonal" and raw) else base)
            cell.font = _font()
            cell.alignment = Alignment(horizontal="left" if key == "title" else "center",
                                       vertical="center", wrap_text=True)
            cell.border = _border_thin()
        ws.row_dimensions[row].height = max(38, _IMG_PX * 0.75 + 2)

    legend_row = 4 + len(products) + 1
    ws.merge_cells(f"A{legend_row}:{last}{legend_row}")
    ws[f"A{legend_row}"] = ("OPPORTUNITY (~115 max) = base 85 (trend/YoY/reviews/rating/weight/price/"
                            "not-Amazon − variation penalty) + revenue-fit 30 (closeness to budget cap)")
    ws[f"A{legend_row}"].font = Font(name="Calibri", bold=True, size=9, color=C["sub_fg"])
    ws[f"A{legend_row}"].fill = _fill(C["sub_bg"])
    for i, (color, text) in enumerate([
        (C["score_hi"], "Opportunity ≥ 80 — strong"),
        (C["score_mid"], "Opportunity 55–79 — moderate"),
        (C["score_lo"], "Opportunity < 55 — weaker"),
        (C["flag_cell"], "Seasonal = Yes — sales concentrated in peak months (review before committing)"),
    ]):
        r = legend_row + 1 + i
        ws.cell(row=r, column=1).fill = _fill(color)
        ws.merge_cells(f"B{r}:{last}{r}")
        ws[f"B{r}"] = text
        ws[f"B{r}"].font = _font(size=9)
        ws.row_dimensions[r].height = 16

    ws.freeze_panes = "A4"


def export_shortlist(
    products: List[Dict],
    batch_num: int,
    *,
    budget: Optional[float] = None,
    category: Optional[str] = None,
    include_seasonal: bool = True,
    pool_total: Optional[int] = None,
) -> str:
    """Write one shortlist batch to a timestamped Excel file in output/."""
    out_dir = Path("output")
    out_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"Phase1_Shortlist_Batch{batch_num}_{ts}.xlsx"

    meta = (
        f"Budget: {(f'£{budget:,.0f}') if budget else 'any'}   |   "
        f"Category: {category or 'any'}   |   "
        f"Seasonal: {'included' if include_seasonal else 'excluded'}   |   "
        f"Pool: {pool_total if pool_total is not None else len(products)}   |   "
        f"Generated: {datetime.now():%d %b %Y %H:%M}"
    )
    wb = openpyxl.Workbook()
    _write_products_sheet(wb, products, batch_num, meta)
    wb.save(path)
    return str(path)


def print_shortlist_table(products: List[Dict], batch_num: int) -> None:
    print(f"\n  ── Shortlist Batch {batch_num} (top {len(products)}) {'─'*30}")
    print(f"  {'#':>2}  {'ASIN':<12}  {'Brand':<16}  {'Price':>7}  {'Sales':>6}  "
          f"{'Rev/mo':>9}  {'Rev':>6}  {'Var':>4}  {'Opp':>5}")
    print(f"  {'─'*92}")
    for i, p in enumerate(products, 1):
        price = _f(p.get("price")); rev = _f(p.get("estimated_revenue"))
        sales = _i(p.get("estimated_sales")); opp = _f(p.get("opportunity_final"))
        price_s = f"£{price:.0f}" if price else "N/A"
        rev_s = f"£{rev:,.0f}" if rev else "N/A"
        sales_s = str(sales) if sales is not None else "N/A"
        opp_s = f"{opp:.1f}" if opp is not None else "N/A"
        print(f"  {i:>2}. {str(p.get('asin','')):<12}  {str(p.get('brand') or '')[:16]:<16}  "
              f"{price_s:>7}  {sales_s:>6}  {rev_s:>9}  {_i(p.get('review_count')) or 0:>6}  "
              f"{_i(p.get('variation_count')) or 0:>4}  {opp_s:>5}")
