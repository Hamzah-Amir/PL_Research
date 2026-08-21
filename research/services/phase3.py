"""
PL Research Tool — Phase 3 (web / library form).

A single, self-contained module for competitor identification + analysis,
flattened from the old research/phase3/ package (auto-assembled). It folds in
the Keepa client, Amazon page-1 scraper, JS-sheet generator, Xray/JS readers,
the 184-point rubric scorer, the Claude listing/review scorers, the FlyBy
reviews client, and the Competitor Analysis workbook writer.

No CLI, no prompts. The high-level entry point `run_phase3(...)` takes its
inputs as parameters and returns data (and writes the .xlsx workbook). Colliding
private helpers from the original modules were prefixed per source module during
assembly; public function names are unchanged.
"""
from __future__ import annotations

from datetime import date
from datetime import datetime
from pathlib import Path
from pydantic import BaseModel, Field
from typing import Dict, List, Optional
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote_plus
import numpy as np
import openpyxl
import os
import pandas as pd
import pickle
import random
import re
import requests
import time

from django.conf import settings
from research.services.phase2 import ProductProfile, derive_product_profile

# Canonical project-root paths (replace each original module's _ROOT/_ENV_PATH).
_ROOT = Path(settings.BASE_DIR)
_ENV_PATH = _ROOT / ".env"
_OUT_DIR = _ROOT / "output"
_ASIN_RE = re.compile(r"^[A-Z0-9]{10}$")


# ==============================================================================
# keepa_client
# ==============================================================================
_CACHE_DIR = _ROOT / "output" / "keepa_cache"

DOMAIN = "GB"  # Amazon UK
_IMG_BASE = "https://m.media-amazon.com/images/I/"


class Phase3KeepaError(RuntimeError):
    """Raised when the Keepa API cannot complete a Phase 3 request."""


# ──────────────────────────────────────────────────────────────────────────────
# ASIN pool
# ──────────────────────────────────────────────────────────────────────────────

def build_asin_pool(
    js_df: Optional[pd.DataFrame],
    xray_dfs: Optional[List[pd.DataFrame]],
) -> List[str]:
    """Deduplicated, order-preserving parent-ASIN pool from JS + all Xray files.

    JS ASINs come first (the workflow's stated source), then any Xray ASINs not
    already present — so a strong competitor surfaced only by Xray is still
    covered. Parents only (variation children come from Keepa later, Phase 4).
    """
    pool: List[str] = []
    seen = set()

    def _add(series):
        for v in series:
            if v is None:
                continue
            s = str(v).strip().upper()
            if _ASIN_RE.match(s) and s not in seen:
                seen.add(s)
                pool.append(s)

    if js_df is not None and "asin" in js_df.columns:
        _add(js_df["asin"])
    for df in (xray_dfs or []):
        if df is not None and "asin" in df.columns:
            _add(df["asin"])
    return pool


# ──────────────────────────────────────────────────────────────────────────────
# Client + disk cache
# ──────────────────────────────────────────────────────────────────────────────

def load_keepa():
    """Build a keepa.Keepa client from the .env KEEPA_API_KEY. No fallback."""
    try:
        from dotenv import load_dotenv
    except ImportError as e:
        raise Phase3KeepaError(
            "python-dotenv is not installed. Run setup again."
        ) from e
    try:
        import keepa  # noqa: F401
    except ImportError as e:
        raise Phase3KeepaError(
            "The 'keepa' package is not installed. Run: pip install keepa."
        ) from e

    load_dotenv(dotenv_path=_ENV_PATH)
    key = os.environ.get("KEEPA_API_KEY")
    if not key:
        raise Phase3KeepaError(
            f"KEEPA_API_KEY not found. Add it to {_ENV_PATH} "
            f"(KEEPA_API_KEY=...). The file is git-ignored."
        )
    import keepa
    try:
        # The library defaults to a 10s read timeout — too short for rich pulls
        # (stats + rating + buybox for a parent + its variations). Give it room.
        return keepa.Keepa(key, timeout=60)
    except Exception as e:  # noqa: BLE001
        raise Phase3KeepaError(f"Could not initialise the Keepa client: {e}") from e


def _cache_path(depth: str) -> Path:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)
    return _CACHE_DIR / f"products_{depth}.pkl"


def _load_cache(depth: str) -> Dict[str, dict]:
    path = _cache_path(depth)
    if path.exists():
        try:
            with open(path, "rb") as fh:
                return pickle.load(fh)
        except Exception:  # noqa: BLE001 — corrupt cache should not be fatal
            return {}
    return {}


def _save_cache(depth: str, cache: Dict[str, dict]) -> None:
    with open(_cache_path(depth), "wb") as fh:
        pickle.dump(cache, fh)


# ──────────────────────────────────────────────────────────────────────────────
# Product query (token-aware, cached)
# ──────────────────────────────────────────────────────────────────────────────

def query_products(
    api,
    asins: List[str],
    *,
    deep: bool = False,
    use_cache: bool = True,
    limit: Optional[int] = None,
) -> Dict[str, dict]:
    """
    Query Keepa for *asins* and return ``{asin: raw_product_dict}``.

    *deep* selects the richer (more expensive) parameter set for the final 3
    competitors. *limit* caps how many *uncached* ASINs are sent to the API in
    this call (token guard — e.g. ``limit=25``). Cached ASINs are always
    returned and never count against the limit.

    Raises Phase3KeepaError on API failure (no fallback).
    """
    depth = "deep" if deep else "shallow"
    cache = _load_cache(depth) if use_cache else {}

    wanted = [a for a in asins if _ASIN_RE.match(a or "")]
    have = {a: cache[a] for a in wanted if a in cache}
    missing = [a for a in wanted if a not in cache]

    if limit is not None and len(missing) > limit:
        missing = missing[:limit]

    if missing:
        kwargs = dict(domain=DOMAIN, history=True, stats=180, progress_bar=False)
        if deep:
            kwargs.update(offers=20, buybox=True, rating=True, aplus=True, videos=True)
        try:
            products = api.query(missing, **kwargs)
        except Exception as e:  # noqa: BLE001
            raise Phase3KeepaError(f"Keepa product query failed: {e}") from e

        for prod in products:
            asin = prod.get("asin")
            if asin:
                cache[asin] = prod
                have[asin] = prod
        if use_cache:
            _save_cache(depth, cache)

    # Preserve the requested order.
    return {a: have[a] for a in wanted if a in have}


def get_seller_feedback(api, seller_id: str) -> Optional[Dict]:
    """All-time seller feedback for *seller_id* via the Keepa Seller endpoint.

    Returns ``{"name", "rating_count", "rating_pct"}`` or None. Used only for
    the final 3 competitors (per the user's all-time-feedback decision).
    """
    if not seller_id or seller_id in ("-1", "-2"):
        return None
    try:
        result = api.seller_query(seller_id, domain=DOMAIN)
    except Exception:  # noqa: BLE001 — seller data is best-effort
        return None
    rec = result.get(seller_id) if isinstance(result, dict) else None
    if not rec:
        return None
    return {
        "name": rec.get("sellerName"),
        # Keepa exposes lifetime counts as currentRatingCount / current rating %.
        "rating_count": rec.get("currentRatingCount") or rec.get("ratingCount"),
        "rating_pct": rec.get("currentRating") or rec.get("rating"),
    }


# ──────────────────────────────────────────────────────────────────────────────
# Field extraction (Keepa-only canonical fields)
# ──────────────────────────────────────────────────────────────────────────────

def _img_urls(product: dict, limit: int = 9) -> List[str]:
    out = []
    for im in (product.get("images") or [])[:limit]:
        name = im.get("l") or im.get("m") if isinstance(im, dict) else None
        if name:
            out.append(_IMG_BASE + name)
    return out


def _kc_listed_since(product: dict) -> Optional[str]:
    raw = product.get("listedSince")
    if not raw or raw <= 0:
        return None
    try:
        import keepa
        return keepa.keepa_minutes_to_time(raw).strftime("%Y-%m-%d")
    except Exception:  # noqa: BLE001
        return None


def extract_keepa_fields(product: dict) -> dict:
    """Pull the canonical Keepa fields Phase 3 needs from a raw product dict.

    Everything here is available in the cheap shallow pull except `a_plus`,
    `has_video`, `buybox_seller_id` and `seller_feedback`, which are populated
    only on the deep (final-3) pull.
    """
    stats = product.get("stats_parsed") or {}
    current = stats.get("current") or {}
    fba = product.get("fbaFees") or {}
    cat_tree = product.get("categoryTree") or []

    pick_pack = fba.get("pickAndPackFee")
    pick_pack = round(pick_pack / 100.0, 2) if isinstance(pick_pack, (int, float)) else None

    variations = product.get("variations") or []

    # A+ / video flags (deep pull). Keepa exposes A+ under aPlusContent when
    # aplus=1 is requested; treat any non-empty value as present.
    aplus_val = product.get("aPlusContent") or product.get("aplus")
    a_plus = bool(aplus_val) if aplus_val not in (None, "", [], {}) else None
    videos = product.get("videos")
    has_video = bool(videos) if videos not in (None, "", [], {}) else None

    bb_hist = product.get("buyBoxSellerIdHistory") or []
    buybox_seller_id = bb_hist[-1] if bb_hist else None

    return {
        "asin": product.get("asin"),
        "parent_asin": product.get("parentAsin"),
        "title": product.get("title"),
        "brand": product.get("brand") or product.get("manufacturer"),
        "price": current.get("NEW") if current.get("NEW", -1) not in (-1, None) else current.get("AMAZON"),
        "list_price": current.get("LISTPRICE") if current.get("LISTPRICE", -1) not in (-1, None) else None,
        "bsr": current.get("SALES") if current.get("SALES", -1) not in (-1, None) else None,
        "bsr_90d": (stats.get("avg90") or {}).get("SALES")
                   if (stats.get("avg90") or {}).get("SALES", -1) not in (-1, None) else None,
        "bullets": list(product.get("features") or []),
        "description": product.get("description"),
        "image_urls": _img_urls(product),
        "images_count": len(product.get("images") or []),
        "variations_count": len(variations),
        "variation_attrs": variations,
        "fba_pick_pack_fee": pick_pack,
        "referral_pct": product.get("referralFeePercent") or product.get("referralFeePercentage"),
        "category": cat_tree[0]["name"] if cat_tree else None,
        "subcategory": cat_tree[-1]["name"] if cat_tree else None,
        "coupon": product.get("coupon"),
        "monthly_sold": product.get("monthlySold"),
        "brand_store_url": product.get("brandStoreUrl"),
        "listed_since": _kc_listed_since(product),
        "item_weight_g": product.get("itemWeight"),
        "package_weight_g": product.get("packageWeight"),
        # deep-only:
        "a_plus": a_plus,
        "has_video": has_video,
        "buybox_seller_id": buybox_seller_id,
        "seller_feedback": None,  # filled by main after get_seller_feedback()
    }


# ==============================================================================
# asin_scraper
# ==============================================================================
try:
    from curl_cffi import requests as _http
    _CURL = True
except ImportError:  # pragma: no cover
    import requests as _http  # type: ignore
    _CURL = False

try:
    from bs4 import BeautifulSoup
except ImportError as _e:  # pragma: no cover
    BeautifulSoup = None  # type: ignore
    _BS4_ERR = _e
else:
    _BS4_ERR = None


_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-GB,en;q=0.9",
    "Upgrade-Insecure-Requests": "1",
}

_CAPTCHA_MARKERS = (
    "api-services-support@amazon.com",
    "Enter the characters you see below",
    "Type the characters you see in this image",
)


def _cookie() -> str:
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=_ENV_PATH)
    except Exception:  # noqa: BLE001
        pass
    return (os.environ.get("AMAZON_COOKIE") or "").strip()


# The free plan only serves the HTTP API endpoint (HTTPS 503s); the target URL
# may still be https. Override with PROXY_ENDPOINT if you upgrade.
_SCRAPESTACK_ENDPOINT = "http://api.scrapestack.com/scrape"
# scrapestack error codes that won't fix themselves on retry (bad key, blocked
# account, plan doesn't allow the feature, monthly quota spent).
_SCRAPESTACK_FATAL = {"101", "102", "103", "104", "105"}


def _scrapestack_key() -> str:
    """Scrapestack access key from .env (PROXY_API_KEY, or SCRAPESTACK_KEY)."""
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=_ENV_PATH)
    except Exception:  # noqa: BLE001
        pass
    return (os.environ.get("PROXY_API_KEY")
            or os.environ.get("SCRAPESTACK_KEY")
            or os.environ.get("SCRAPESTACK_ACCESS_KEY") or "").strip()


def _scrapestack_error(text: str) -> Optional[str]:
    """If *text* is a scrapestack JSON error body, return 'code type', else None.
    (apilayer APIs return HTTP 200 with {"success": false, "error": {...}}.)"""
    t = (text or "").lstrip()
    if not t.startswith("{") or "\"success\"" not in t:
        return None
    try:
        import json
        data = json.loads(t)
    except Exception:  # noqa: BLE001
        return None
    if isinstance(data, dict) and data.get("success") is False:
        err = data.get("error") or {}
        return f"{err.get('code')} {err.get('type')}".strip()
    return None


def _asc_get(url: str, cookie: str, timeout: float = 30):
    """Fetch *url*. When a scrapestack key is configured (PROXY_API_KEY), route
    the request through the scrapestack proxy — its rotating 35M-IP pool gets
    past Amazon's bot-blocking. Otherwise hit Amazon directly (curl_cffi)."""
    key = _scrapestack_key()
    if key:
        params = {"access_key": key, "url": url}
        # Optional, off by default (render_js / premium_proxy need a paid plan;
        # premium counts as 25 API calls). Enable via .env if your plan allows.
        if (os.environ.get("PROXY_RENDER_JS") or "").strip() not in ("", "0"):
            params["render_js"] = "1"
        if (os.environ.get("PROXY_PREMIUM") or "").strip() not in ("", "0"):
            params["premium_proxy"] = "1"
        loc = (os.environ.get("PROXY_LOCATION") or "").strip()
        if loc:
            params["proxy_location"] = loc
        headers = {}
        if cookie:  # forward the Amazon session cookie to the target page
            params["keep_headers"] = "1"
            headers = {"Cookie": cookie, "User-Agent": _HEADERS["User-Agent"]}
        # Proxy round-trips (esp. with render_js/premium) are slower than direct.
        endpoint = (os.environ.get("PROXY_ENDPOINT") or "").strip() or _SCRAPESTACK_ENDPOINT
        return requests.get(endpoint, params=params, headers=headers,
                            timeout=max(timeout, 90))
    headers = dict(_HEADERS)
    if cookie:
        headers["Cookie"] = cookie
    if _CURL:
        return _http.get(url, headers=headers, timeout=timeout, impersonate="chrome")
    return _http.get(url, headers=headers, timeout=timeout)


def _is_captcha(html: str) -> bool:
    return any(m in html for m in _CAPTCHA_MARKERS)


def _looks_blocked(html: str) -> bool:
    """Amazon intermittently serves a tiny no-JS shell to bots: HTTP 200, no
    CAPTCHA text, but no product grid. Real result pages (even a genuine
    "no results" page) are hundreds of KB and contain data-asin attributes, so a
    short page with no data-asin is a soft block worth retrying."""
    return len(html) < 100_000 and "data-asin" not in html


def _looks_challenge(html: str) -> bool:
    """Amazon's bot-mitigation interstitial: a tiny HTML page with a meta-refresh
    carrying a `bm-`/`bm/` token instead of the results. Only a JS-rendering or
    residential proxy can clear it, so retrying a plain datacenter proxy just
    wastes quota."""
    if len(html) > 20_000:
        return False
    low = html.lower()
    return 'http-equiv="refresh"' in low and ("bm-" in html or "/bm/" in low or "botmitigation" in low)


def _fetch_page(url: str, cookie: str, retries: int = 4) -> Optional[str]:
    """Fetch one results page; retry on CAPTCHA / soft-block / 5xx with backoff.
    None = blocked after all retries."""
    proxy = bool(_scrapestack_key())
    render_js = proxy and (os.environ.get("PROXY_RENDER_JS") or "").strip() not in ("", "0")
    if proxy:
        retries = min(retries, 2)  # each attempt spends a scrapestack request
    for attempt in range(retries):
        try:
            r = _asc_get(url, cookie)
        except Exception:  # noqa: BLE001 — network hiccup, retry
            time.sleep(2.0 * (attempt + 1))
            continue
        # Proxy-layer error (bad key / quota / plan). Fatal ones won't fix on
        # retry — bail immediately so the caller falls back cleanly.
        serr = _scrapestack_error(r.text)
        if serr:
            print(f"  scrapestack error: {serr} for {url}")
            if serr.split(" ", 1)[0] in _SCRAPESTACK_FATAL:
                return None
            time.sleep(3.0 * (attempt + 1) + random.uniform(0, 2))
            continue
        # Amazon bot-mitigation interstitial. Without JS rendering a datacenter
        # proxy can't clear it — stop now to conserve the request quota.
        if not render_js and _looks_challenge(r.text):
            print("  Amazon bot-mitigation challenge via proxy — needs render_js "
                  "(paid plan) or premium/residential proxy. Set PROXY_RENDER_JS=1 "
                  "once your scrapestack plan supports it.")
            return None
        if r.status_code == 200 and not _is_captcha(r.text):
            if _looks_blocked(r.text):
                time.sleep(3.0 * (attempt + 1) + random.uniform(0, 2))
                continue  # soft-block shell — retry for the full page
            return r.text
        if r.status_code == 404:
            return None
        time.sleep(3.0 * (attempt + 1) + random.uniform(0, 2))
    return None


def _is_sponsored(block) -> bool:
    if block.select_one(
        '[data-component-type="sp-sponsored-result"], '
        ".puis-sponsored-label-text, .s-sponsored-label-info-icon"
    ):
        return True
    for el in block.select("a[aria-label], span[aria-label]"):
        if "sponsored" in (el.get("aria-label") or "").lower():
            return True
    return False


def _parse_results(html: str) -> List[Dict]:
    """Extract {asin, title, sponsored} blocks from one search page, in order.

    Amazon's result markup varies (logged-in vs anonymous, A/B layouts), so we
    fall back through progressively broader selectors instead of relying on a
    single one: the canonical search-result component, then any result item
    carrying a data-asin, then any element with a valid data-asin.
    """
    soup = BeautifulSoup(html, "html.parser")
    blocks = (soup.select('div[data-component-type="s-search-result"]')
              or soup.select('div.s-result-item[data-asin]')
              or [el for el in soup.select('[data-asin]')
                  if _ASIN_RE.fullmatch((el.get("data-asin") or "").strip().upper())])

    out, seen = [], set()
    for block in blocks:
        asin = (block.get("data-asin") or "").strip().upper()
        if not _ASIN_RE.fullmatch(asin) or asin in seen:
            continue
        seen.add(asin)
        title_el = (block.select_one("h2 a span") or block.select_one("h2 span")
                    or block.select_one("h2")
                    or block.select_one('[data-cy="title-recipe"] span')
                    or block.select_one("a.s-line-clamp-2, a.s-line-clamp-3, a.s-line-clamp-4"))
        out.append({
            "asin": asin,
            "title": title_el.get_text(" ", strip=True) if title_el else "",
            "sponsored": _is_sponsored(block),
        })
    return out


def _has_next_page(html: str) -> bool:
    soup = BeautifulSoup(html, "html.parser")
    nxt = soup.select_one("a.s-pagination-next")
    if nxt is not None:
        return True
    # Disabled "Next" renders as a span, an explicit signal of the last page.
    return False


def _search_url(keyword: str, marketplace: str, page: int) -> str:
    base = f"https://www.amazon.{marketplace}/s?k={quote_plus(keyword)}"
    return f"{base}&page={page}" if page > 1 else base


def _page_url(url: str, page: int) -> str:
    stripped = re.sub(r"([&?])page=\d+", r"\1", url).rstrip("&?")
    if page == 1:
        return stripped
    sep = "&" if "?" in stripped else "?"
    return f"{stripped}{sep}page={page}"


def scrape_asins(
    keyword: Optional[str] = None,
    url: Optional[str] = None,
    marketplace: str = "co.uk",
    max_pages: int = 1,
    include_sponsored: bool = True,
    delay: float = 1.5,
) -> Dict:
    """Scrape every ASIN from Amazon search results for `keyword` (or a full
    search/listing `url`), walking pages 1..max_pages in ranked order
    (default: page 1 only).

    Returns {keyword, url, marketplace, pages_fetched, pages_blocked, count,
    results[{asin, title, page, position, sponsored}], note}. Duplicates across
    pages keep their first (highest-ranked) occurrence. Never raises for normal
    failures — reports via note.
    """
    if BeautifulSoup is None:
        return {"error": f"beautifulsoup4 not installed ({_BS4_ERR})", "results": [], "count": 0}
    if not keyword and not url:
        return {"error": "Provide a keyword or a url.", "results": [], "count": 0}

    cookie = _cookie()
    results: List[Dict] = []
    seen: set = set()
    pages_fetched, pages_blocked = 0, 0
    position = 0

    for page in range(1, max_pages + 1):
        page_url = _page_url(url, page) if url else _search_url(keyword, marketplace, page)
        html = _fetch_page(page_url, cookie)
        if html is None:
            pages_blocked += 1
            break
        pages_fetched += 1
        blocks = _parse_results(html)
        if not blocks:
            break
        for b in blocks:
            if b["asin"] in seen:
                continue
            if not include_sponsored and b["sponsored"]:
                continue
            seen.add(b["asin"])
            position += 1
            results.append({**b, "page": page, "position": position})
        if not _has_next_page(html):
            break
        time.sleep(delay + random.uniform(0, 1))

    note = ""
    if pages_blocked:
        note = (
            f"Page {pages_fetched + 1} was blocked (CAPTCHA or repeated errors) — "
            f"results cover the first {pages_fetched} page(s) only."
        )
    elif pages_fetched == 0:
        note = "No pages could be fetched."
    elif not results:
        note = "Pages fetched but no search results found — selector mismatch or empty search."

    return {
        "keyword": keyword,
        "url": url,
        "marketplace": marketplace,
        "pages_fetched": pages_fetched,
        "pages_blocked": pages_blocked,
        "count": len(results),
        "sponsored_count": sum(1 for r in results if r["sponsored"]),
        "results": results,
        "note": note,
        "scraper": "curl_cffi" if _CURL else "requests",
    }


def asin_list(result: Dict, organic_only: bool = False) -> List[str]:
    """Flatten a scrape result into the ordered ASIN list consumed by the
    Keepa stage (empty if nothing usable)."""
    rows = result.get("results") or []
    if organic_only:
        rows = [r for r in rows if not r.get("sponsored")]
    return [r["asin"] for r in rows]


# ==============================================================================
# js_sheet
# ==============================================================================
# Final sheet columns — the JS export's columns, plus three honest extras.
COLUMNS = [
    "Position", "ASIN", "Product Name", "Brand", "Price", "Monthly Units Sold",
    "Daily Units Sold", "Monthly Revenue", "Date First Available", "Net Revenue",
    "Star Rating", "Reviews", "Amazon Fees", "BSR", "LQS (est)", "Fulfillment",
    "No. of Sellers", "Category", "Product Tier", "Dimensions", "Weight (kg)",
    "Link", "Sponsored", "Sales Source",
]

# Amazon UK FBA size tiers: (name, max_kg, dims_cm sorted descending).
_UK_TIERS = [
    ("Light envelope", 0.10, (33, 23, 2.5)),
    ("Standard envelope", 0.46, (33, 23, 2.5)),
    ("Large envelope", 0.96, (33, 23, 4)),
    ("Extra-Large envelope", 0.96, (33, 23, 6)),
    ("Small parcel", 3.9, (35, 25, 12)),
    ("Standard parcel", 11.9, (45, 34, 26)),
]


def uk_size_tier(l_cm, w_cm, h_cm, kg) -> Optional[str]:
    """Amazon UK size tier from package dims (cm) + weight (kg). None if unknown."""
    if not all(isinstance(v, (int, float)) and v > 0 for v in (l_cm, w_cm, h_cm, kg)):
        return None
    dims = sorted((l_cm, w_cm, h_cm), reverse=True)
    for name, max_kg, lims in _UK_TIERS:
        if kg <= max_kg and all(d <= m for d, m in zip(dims, lims)):
            return name
    return "Oversize"


def _lqs_estimate(images, bullets, title, description, rating, reviews) -> int:
    """Our own 1-10 listing-quality score (JS's LQS is proprietary)."""
    pts = 0
    pts += 3 if images >= 7 else 2 if images >= 5 else 1 if images >= 1 else 0
    pts += 2 if bullets >= 5 else 1 if bullets >= 3 else 0
    tl = len(title or "")
    pts += 2 if tl >= 80 else 1 if tl >= 40 else 0
    pts += 1 if (description or "").strip() else 0
    pts += 1 if isinstance(rating, (int, float)) and rating >= 4.3 else 0
    pts += 1 if isinstance(reviews, (int, float)) and reviews >= 100 else 0
    return min(pts, 10)


def _cur(current: dict, key: str):
    v = current.get(key)
    return v if isinstance(v, (int, float)) and v != -1 else None


def _jss_listed_since(product: dict) -> Optional[str]:
    raw = product.get("listedSince")
    if not isinstance(raw, (int, float)) or raw <= 0:
        return None
    try:
        import keepa
        return keepa.keepa_minutes_to_time(raw).strftime("%Y-%m-%d")
    except Exception:  # noqa: BLE001
        return None


def _fulfillment(product: dict, stats: dict, current: dict) -> Optional[str]:
    bb_fba = stats.get("buyBoxIsFBA")  # only present with buybox=True
    if product.get("availabilityAmazon") == 0:
        return "AMZ"
    if bb_fba is True:
        return "FBA"
    if bb_fba is False:
        return "FBM"
    # Without buy-box data: a current FBA/FBM offer price is the next-best signal.
    if _cur(current, "NEW_FBA") is not None:
        return "FBA"
    if _cur(current, "NEW_FBM_SHIPPING") is not None:
        return "FBM"
    return None


def product_to_row(product: dict, marketplace: str = "co.uk") -> Dict:
    """Derive one JS-sheet row from a raw Keepa product dict (blanks = unknown)."""
    stats = product.get("stats_parsed") or {}
    current = stats.get("current") or {}

    price = _cur(current, "NEW")
    if price is None:
        price = _cur(current, "AMAZON")
    if price is None:
        price = _cur(current, "BUY_BOX_SHIPPING")

    monthly = product.get("monthlySold")
    source = "Amazon badge" if isinstance(monthly, (int, float)) and monthly > 0 else None
    if source is None:
        drops = stats.get("salesRankDrops30")
        if isinstance(drops, (int, float)) and drops > 0:
            monthly, source = drops, "BSR drops (est.)"
        else:
            monthly = None

    pick_pack = (product.get("fbaFees") or {}).get("pickAndPackFee")
    pick_pack = pick_pack / 100.0 if isinstance(pick_pack, (int, float)) else None
    ref_pct = product.get("referralFeePercentage") or product.get("referralFeePercent")
    fees = None
    if pick_pack is not None and isinstance(ref_pct, (int, float)) and price is not None:
        fees = round(pick_pack + price * ref_pct / 100.0, 2)

    rating = _cur(current, "RATING")
    reviews = _cur(current, "COUNT_REVIEWS")

    dims_mm = [product.get(k) for k in ("packageLength", "packageWidth", "packageHeight")]
    dims_cm = [round(v / 10.0, 1) if isinstance(v, (int, float)) and v > 0 else None for v in dims_mm]
    weight_g = product.get("packageWeight") or product.get("itemWeight")
    kg = round(weight_g / 1000.0, 2) if isinstance(weight_g, (int, float)) and weight_g > 0 else None

    cat_tree = product.get("categoryTree") or []
    asin = product.get("asin")
    title = product.get("title")

    return {
        "ASIN": asin,
        "Product Name": title,
        "Brand": product.get("brand") or product.get("manufacturer"),
        "Price": price,
        "Monthly Units Sold": monthly,
        "Daily Units Sold": round(monthly / 30.0, 1) if monthly is not None else None,
        "Monthly Revenue": round(monthly * price, 2) if monthly is not None and price is not None else None,
        "Date First Available": _jss_listed_since(product),
        "Net Revenue": round(price - fees, 2) if price is not None and fees is not None else None,
        "Star Rating": rating,
        "Reviews": reviews,
        "Amazon Fees": fees,
        "BSR": _cur(current, "SALES"),
        "LQS (est)": _lqs_estimate(
            len(product.get("images") or []), len(product.get("features") or []),
            title, product.get("description"), rating, reviews,
        ),
        "Fulfillment": _fulfillment(product, stats, current),
        "No. of Sellers": _cur(current, "COUNT_NEW"),
        "Category": cat_tree[0]["name"] if cat_tree else None,
        "Product Tier": uk_size_tier(*dims_cm, kg),
        "Dimensions": (" x ".join(str(d) for d in dims_cm) + " cm") if all(d is not None for d in dims_cm) else None,
        "Weight (kg)": kg,
        "Link": f"https://www.amazon.{marketplace}/dp/{asin}",
        "Sales Source": source,
    }


def _query_sheet(api, asins: List[str], *, buybox: bool, use_cache: bool,
                 limit: Optional[int]):
    """Batch Keepa Product Request for the sheet (history+stats+rating).

    Cached per ASIN at depth "sheet"; missing ASINs are borrowed from the
    phase-3 deep cache when present (its params are a superset). `limit` caps
    uncached ASINs sent to the API.
    """
    cache = _load_cache("sheet") if use_cache else {}
    deep = _load_cache("deep") if use_cache else {}

    wanted = [a for a in asins if _ASIN_RE.fullmatch((a or "").strip().upper())]
    have, missing = {}, []
    for a in wanted:
        if a in cache:
            have[a] = cache[a]
        elif a in deep:
            have[a] = cache[a] = deep[a]
        else:
            missing.append(a)

    skipped = 0
    if limit is not None and len(missing) > limit:
        skipped = len(missing) - limit
        missing = missing[:limit]

    if missing:
        kwargs = dict(domain=DOMAIN, history=True, stats=180, rating=True, progress_bar=False)
        if buybox:
            kwargs.update(buybox=True)
        for i in range(0, len(missing), 100):  # endpoint cap: 100 ASINs/call
            try:
                products = api.query(missing[i:i + 100], **kwargs)
            except Exception as e:  # noqa: BLE001
                raise Phase3KeepaError(f"Keepa product query failed: {e}") from e
            for prod in products:
                if prod.get("asin"):
                    cache[prod["asin"]] = have[prod["asin"]] = prod
        if use_cache:
            _save_cache("sheet", cache)

    return {a: have[a] for a in wanted if a in have}, skipped


def build_js_sheet(
    keyword: Optional[str] = None,
    asins: Optional[List[str]] = None,
    *,
    marketplace: str = "co.uk",
    max_pages: int = 1,
    out_path: Optional[str] = None,
    buybox: bool = False,
    use_cache: bool = True,
    limit: Optional[int] = None,
) -> Dict:
    """Keyword (or explicit ASIN list) -> Keepa -> JS-style Excel sheet.

    Returns {out_path, count, skipped, tokens_left, summary, note}.
    """
    import pandas as pd

    scrape_meta = {}
    if asins:
        # Accept plain ASIN strings or scraper-result dicts ({asin, sponsored, position}).
        rows_meta = []
        for i, a in enumerate(asins):
            if isinstance(a, dict):
                rows_meta.append({"asin": (a.get("asin") or "").strip().upper(),
                                  "sponsored": a.get("sponsored"),
                                  "position": a.get("position") or i + 1})
            else:
                rows_meta.append({"asin": a.strip().upper(), "sponsored": None, "position": i + 1})
    else:
        if not keyword:
            raise ValueError("Provide a keyword or an ASIN list.")
        scraped = scrape_asins(keyword=keyword, marketplace=marketplace, max_pages=max_pages)
        if scraped.get("error") or not scraped.get("results"):
            raise Phase3KeepaError(
                f"ASIN scrape failed: {scraped.get('error') or scraped.get('note') or 'no results'}"
            )
        rows_meta = scraped["results"]
        scrape_meta = {"pages_fetched": scraped["pages_fetched"], "note": scraped.get("note", "")}

    api = load_keepa()
    products, skipped = _query_sheet(
        api, [r["asin"] for r in rows_meta], buybox=buybox, use_cache=use_cache, limit=limit,
    )

    rows = []
    for meta in rows_meta:
        prod = products.get(meta["asin"])
        if prod is None:
            continue  # not pulled (token limit) or unknown to Keepa
        row = product_to_row(prod, marketplace)
        row["Position"] = meta.get("position")
        row["Sponsored"] = meta.get("sponsored")
        rows.append(row)

    df = pd.DataFrame(rows, columns=COLUMNS)

    def _avg(col):
        s = pd.to_numeric(df[col], errors="coerce").dropna()
        return round(float(s.mean()), 2) if len(s) else None

    summary = {
        "Keyword": keyword or "(explicit ASIN list)",
        "Marketplace": f"amazon.{marketplace}",
        "Export date": date.today().isoformat(),
        "Products": len(df),
        "Avg Monthly Sales": _avg("Monthly Units Sold"),
        "Avg Monthly Revenue": _avg("Monthly Revenue"),
        "Avg Price": _avg("Price"),
        "Avg Reviews": _avg("Reviews"),
        "Avg BSR": _avg("BSR"),
        "Avg Rating": _avg("Star Rating"),
    }

    if out_path is None:
        slug = re.sub(r"[^a-z0-9]+", "_", (keyword or "asins").lower()).strip("_")
        _OUT_DIR.mkdir(parents=True, exist_ok=True)
        out_path = str(_OUT_DIR / f"js_sheet_{slug}_{date.today().isoformat()}.csv")

    # utf-8-sig so Excel opens the £/unicode columns correctly.
    df.to_csv(out_path, index=False, encoding="utf-8-sig")

    note = scrape_meta.get("note", "")
    if skipped:
        note = (note + " " if note else "") + (
            f"{skipped} ASIN(s) skipped by --limit token guard (re-run to fetch; cache keeps what's paid for)."
        )
    return {
        "out_path": out_path,
        "count": len(df),
        "skipped": skipped,
        "tokens_left": getattr(api, "tokens_left", None),
        "summary": summary,
        "note": note,
    }


def start_background(keyword: str):
    """Kick off `build_js_sheet(keyword)` in a daemon thread.

    Called by Phase 2 the moment Claude extracts the main search term from the
    target's title, so the sheet generates while the user works through the
    keyword steps. Returns ``(thread, holder)``; after ``thread.join()``,
    ``holder`` has "res" (build_js_sheet result) or "err" (the exception).
    """
    import threading

    holder: Dict = {}

    def _gen():
        try:
            holder["res"] = build_js_sheet(keyword=keyword)
        except Exception as e:  # noqa: BLE001 — surfaced by the consumer after join()
            holder["err"] = e

    t = threading.Thread(target=_gen, daemon=True)
    t.start()
    return t, holder


# ==============================================================================
# scorer
# ==============================================================================
MAX_TOTAL = 100          # the raw 184-point rubric is rescaled to /100 for output
_RAW_MAX_TOTAL = 184
BANDS = [
    (0, 47, "Easy to compete — high chance of success for a new seller"),
    (48, 73, "Mid-range challenge — need a strong plan to beat weaknesses"),
    (74, 100, "Difficult — established competition, high chance of failure"),
]

# Element order + max points. Marketing title/bullets reuse the product-page
# judgement (same listing); pricing strategy is Claude's.
ELEMENTS = [
    # key, label, max, section
    ("price", "Price", 25, "PRODUCT PAGE"),
    ("sponsored_products", "Sponsored Products", 20, "PRODUCT PAGE"),
    ("sponsored_brands", "Sponsored Brands (ex-AMS)", 10, "PRODUCT PAGE"),
    ("reviews", "# of Product Reviews", 10, "PRODUCT PAGE"),
    ("rating", "Avg. Product Rating", 5, "PRODUCT PAGE"),
    ("age", "Age of Product", 5, "PRODUCT PAGE"),
    ("product_images", "Product Images", 20, "PRODUCT PAGE"),
    ("product_title", "Product Title", 5, "PRODUCT PAGE"),
    ("bullets_description", "Product Description/Bullet Points", 5, "PRODUCT PAGE"),
    ("bestseller", "Bestseller Badge", 10, "PRODUCT PAGE"),
    ("seller", "Amazon or 3P Seller", 5, "PRODUCT PAGE"),
    ("unique_design", "Unique Product Design", 2, "PRODUCT PAGE"),
    ("fba_fbm", "FBA or FBM", 2, "PRODUCT PAGE"),
    ("mkt_images", "Marketing — Product Images", 15, "MARKETING STRATEGY"),
    ("mkt_title", "Marketing — Product Title", 5, "MARKETING STRATEGY"),
    ("mkt_bullets", "Marketing — Description/Bullets", 5, "MARKETING STRATEGY"),
    ("pricing_strategy", "Pricing Strategy", 10, "MARKETING STRATEGY"),
    ("review_velocity", "Review Velocity", 5, "MARKET MOMENTUM"),
    ("variations", "Product Variations", 5, "MARKET MOMENTUM"),
    ("sales_strength", "Sales / Revenue Strength", 10, "MARKET MOMENTUM"),
    ("enhanced_content", "Enhanced Content (A+/Video)", 5, "MARKET MOMENTUM"),
]
ELEMENT_MAX = {k: m for k, _, m, _ in ELEMENTS}


def _rescale_to_100(elements):
    """Proportionally rescale the raw element maxes (sum 184) to integer weights
    that sum to exactly 100 (largest-remainder method), preserving relative
    weightage."""
    raw = [(k, m) for k, _, m, _ in elements]
    total = sum(m for _, m in raw) or 1
    exact = [(k, m * 100.0 / total) for k, m in raw]
    weights = {k: int(x) for k, x in exact}
    for _, k in sorted(((x - int(x), k) for k, x in exact), reverse=True)[:100 - sum(weights.values())]:
        weights[k] += 1
    return weights


# Per-element /100 weight (sums to exactly 100). Raw ELEMENT_MAX still drives the
# scoring math; scores are rescaled to these weights for the output.
SCALED_ELEMENT_MAX = _rescale_to_100(ELEMENTS)


def _scr_num(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() in ("nan", "n/a", "-"):
        return None
    # JS / Xray numeric dates are US mm/dd/yyyy (e.g. 12/14/2025, 03/30/2023),
    # so try that before dd/mm/yyyy (which still catches true UK-format dates).
    for fmt in ("%m/%d/%Y", "%d/%m/%Y", "%b %d, %Y", "%B %d, %Y", "%Y-%m-%d", "%d %b %Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _percentile(value: Optional[float], pool: List[float]) -> Optional[float]:
    """Fraction of *pool* values <= *value* (0..1). None if no data."""
    vals = [v for v in pool if v is not None]
    if value is None or not vals:
        return None
    return sum(1 for v in vals if v <= value) / len(vals)


def _rel_low_is_strong(value, values) -> Optional[float]:
    """0..1 where the LOWEST value scores 1.0 (used for price)."""
    vals = [v for v in values if v is not None]
    if value is None or len(vals) < 2:
        return None if value is None else 0.5
    lo, hi = min(vals), max(vals)
    if hi == lo:
        return 0.5
    return 1.0 - (value - lo) / (hi - lo)


def _e(score, mx, reason, info=""):
    """Build an element result; score None = Unknown (not counted)."""
    if score is not None:
        score = max(0, min(int(round(score)), mx))
    return {"score": score, "max": mx, "reason": reason, "info": info}


def score_competitor(
    comp: Dict,
    peers: List[Dict],
    pool_stats: Dict,
    listing,
) -> Dict:
    """Score one competitor across all 21 elements.

    *comp* — merged record + Keepa deep fields + sponsored counts + review_velocity.
    *peers* — the selected competitor set (for relative price/sales among them).
    *pool_stats* — {'review_counts': [...], 'sales': [...]} from the same-type
        pool (for category-relative reviews / sales strength).
    *listing* — the Claude `ListingScores` for this competitor (judgement elements).
    """
    el: Dict[str, Dict] = {}
    peer_prices = [_scr_num(p.get("price")) for p in peers]
    peer_sales = [_scr_num(p.get("monthly_sales")) for p in peers]
    total_kw = comp.get("total_keywords") or 0

    # 1. Price (lower = stronger)
    pos = _rel_low_is_strong(_scr_num(comp.get("price")), peer_prices)
    if pos is None:
        el["price"] = _e(None, 25, "No price available.", "Unknown — need user confirmation")
    else:
        el["price"] = _e(25 * pos, 25, f"Price £{comp.get('price')} vs peers (lower=stronger).",
                         f"£{comp.get('price')}")

    # 2. Sponsored Products (presence across launch keywords)
    spc = comp.get("sp_count") or 0
    if total_kw:
        el["sponsored_products"] = _e(20 * spc / total_kw, 20,
            f"Sponsored Products on {spc}/{total_kw} launch keywords.",
            f"SP on {spc}/{total_kw} kws")
    else:
        el["sponsored_products"] = _e(None, 20, "No Xray keyword data.", "Unknown — need user confirmation")

    # 3. Sponsored Brands (ex-AMS)
    sbc = comp.get("sb_count") or 0
    if total_kw:
        el["sponsored_brands"] = _e(10 * sbc / total_kw, 10,
            f"Sponsored Brand ads on {sbc}/{total_kw} launch keywords.",
            f"SB on {sbc}/{total_kw} kws")
    else:
        el["sponsored_brands"] = _e(None, 10, "No Xray keyword data.", "Unknown — need user confirmation")

    # 4. Reviews (category-relative percentile)
    rc = _scr_num(comp.get("review_count"))
    pct = _percentile(rc, pool_stats.get("review_counts", []))
    if pct is None:
        el["reviews"] = _e(None, 10, "No review count.", "Unknown — need user confirmation")
    else:
        el["reviews"] = _e(10 * pct, 10,
            f"{int(rc):,} reviews — {pct*100:.0f}th percentile in category.", f"{int(rc):,}")

    # 5. Rating
    rt = _scr_num(comp.get("rating"))
    if rt is None:
        el["rating"] = _e(None, 5, "No rating.", "Unknown — need user confirmation")
    else:
        sc = 5 if rt >= 4.7 else 4 if rt >= 4.5 else 3 if rt >= 4.3 else 2 if rt >= 4.0 else 1 if rt >= 3.5 else 0
        el["rating"] = _e(sc, 5, f"Rating {rt}.", f"{rt}")

    # 6. Age of product (Date First Available / Creation Date)
    dt = _parse_date(comp.get("date_first_available"))
    if dt is None:
        el["age"] = _e(None, 5, "No reliable date first available.", "Unknown — need user confirmation")
    else:
        yrs = (datetime.now() - dt).days / 365.25
        sc = 5 if yrs >= 3 else 4 if yrs >= 2 else 3 if yrs >= 1.5 else 2 if yrs >= 1 else 1 if yrs >= 0.5 else 0
        el["age"] = _e(sc, 5, f"~{yrs:.1f} years old (older=more established).", dt.strftime("%Y-%m-%d"))

    # 7-9, 12, 14-17. Judgement elements from Claude
    def _claude(attr, mx, reason_attr):
        if listing is None:
            return _e(None, mx, "Claude listing scoring unavailable.", "Unknown — need user confirmation")
        val = getattr(listing, attr, None)
        if val is None or val < 0:  # -1 sentinel = cannot judge
            return _e(None, mx, getattr(listing, reason_attr, ""), "Unknown — need user confirmation")
        return _e(val, mx, getattr(listing, reason_attr, ""))

    el["product_images"] = _claude("product_images", 20, "product_images_reason")
    el["product_title"] = _claude("product_title", 5, "product_title_reason")
    el["bullets_description"] = _claude("bullets_description", 5, "bullets_description_reason")
    el["unique_design"] = _claude("unique_design", 2, "unique_design_reason")
    el["mkt_images"] = _claude("marketing_images", 15, "marketing_images_reason")
    # Marketing title/bullets reuse the product-page judgement (same listing).
    el["mkt_title"] = _claude("product_title", 5, "product_title_reason")
    el["mkt_bullets"] = _claude("bullets_description", 5, "bullets_description_reason")
    el["pricing_strategy"] = _claude("pricing_strategy", 10, "pricing_strategy_reason")

    # 10. Bestseller badge (flag + BSR proxy)
    bs = str(comp.get("best_seller") or "").strip().lower()
    bsr = _scr_num(comp.get("bsr"))
    if bs in ("yes", "true", "1"):
        el["bestseller"] = _e(9, 10, "Holds a Best Seller badge.", "Best Seller: Yes")
    elif bsr is not None:
        sc = 5 if bsr <= 20 else 3 if bsr <= 100 else 2 if bsr <= 500 else 1
        el["bestseller"] = _e(sc, 10, f"No badge; BSR #{int(bsr)} (rank proxy).", f"No badge / BSR #{int(bsr)}")
    else:
        el["bestseller"] = _e(None, 10, "No badge/BSR data.", "Unknown — need user confirmation")

    # 11. Amazon or 3P seller (all-time feedback, deep pull)
    fb = comp.get("seller_feedback") or {}
    fb_count = _scr_num(fb.get("rating_count"))
    if fb_count is not None:
        sc = 5 if fb_count >= 10000 else 4 if fb_count >= 2000 else 3 if fb_count >= 500 else 2 if fb_count >= 100 else 1
        el["seller"] = _e(sc, 5, f"Seller all-time feedback {int(fb_count):,}.", f"{int(fb_count):,} feedback")
    else:
        el["seller"] = _e(None, 5, "All-time seller feedback unavailable (needs offers pull / 3P buy box).",
                          "Unknown — need user confirmation")

    # 13. FBA or FBM
    ff = str(comp.get("fulfillment") or "").upper()
    if ff in ("FBA", "AMZ"):
        el["fba_fbm"] = _e(2, 2, "Fulfilled by Amazon (FBA/AMZ).", ff)
    elif ff == "FBM":
        el["fba_fbm"] = _e(0, 2, "Fulfilled by Merchant.", "FBM")
    else:
        el["fba_fbm"] = _e(None, 2, "Unknown fulfilment.", "Unknown — need user confirmation")

    # 18. Review velocity
    rv = _scr_num(comp.get("review_velocity"))
    if rv is None:
        el["review_velocity"] = _e(None, 5, "No review-velocity data.", "Unknown — need user confirmation")
    else:
        sc = 5 if rv >= 30 else 3 if rv >= 10 else 1 if rv > 0 else 0
        el["review_velocity"] = _e(sc, 5, f"~{rv:.0f} reviews/month.", f"{rv:.0f}/mo")

    # 19. Product variations
    vc = _scr_num(comp.get("variations_count"))
    if vc is None:
        el["variations"] = _e(None, 5, "No variation data.", "Unknown — need user confirmation")
    else:
        sc = 5 if vc >= 5 else 3 if vc >= 2 else 1 if vc >= 1 else 0
        el["variations"] = _e(sc, 5, f"{int(vc)} variation(s).", f"{int(vc)}")

    # 20. Sales / revenue strength (category-relative)
    sp = _percentile(_scr_num(comp.get("monthly_sales")), pool_stats.get("sales", []))
    if sp is None:
        el["sales_strength"] = _e(None, 10, "No sales data.", "Unknown — need user confirmation")
    else:
        ms = _scr_num(comp.get("monthly_sales"))
        el["sales_strength"] = _e(10 * sp, 10,
            f"{int(ms):,} units/mo — {sp*100:.0f}th percentile in category.", f"{int(ms):,}/mo")

    # 21. Enhanced content (A+ / video / image count)
    ap = comp.get("a_plus")
    vid = comp.get("has_video")
    imgs = _scr_num(comp.get("images_count"))
    if ap is None and vid is None and imgs is None:
        el["enhanced_content"] = _e(None, 5, "No content data.", "Unknown — need user confirmation")
    else:
        if ap and vid:
            sc, why = 5, "A+ content and video present."
        elif ap or vid:
            sc, why = 4, "A+ content or video present."
        elif imgs is not None and imgs >= 6:
            sc, why = 3, f"{int(imgs)} images, no A+/video detected."
        else:
            sc, why = 1, "Minimal media."
        el["enhanced_content"] = _e(sc, 5, why, f"A+={ap} video={vid} imgs={imgs}")

    # Rescale the raw 184-point rubric to a 100-point scale (proportional
    # weightage): each element's max becomes its /100 weight and its score is
    # scaled to that weight. Relative weightings are preserved; the total is now
    # out of 100.
    for k, v in el.items():
        raw_max = ELEMENT_MAX.get(k) or v.get("max")
        new_max = SCALED_ELEMENT_MAX.get(k, raw_max)
        if v.get("score") is not None and raw_max:
            v["score"] = round(v["score"] * new_max / raw_max)
        v["max"] = new_max

    # Totals (now on the /100 scale)
    total = sum(v["score"] for v in el.values() if v["score"] is not None)
    unknown = [k for k, v in el.items() if v["score"] is None]
    band = next((d for lo, hi, d in BANDS if lo <= total <= hi), "")

    return {
        "asin": comp.get("asin"),
        "elements": el,
        "total": total,
        "max": MAX_TOTAL,
        "band": band,
        "unknown_elements": unknown,
    }


def pool_statistics(same_type_records) -> Dict:
    """Build {'review_counts','sales'} distributions from the same-type pool
    (a DataFrame or list of dicts) for category-relative scoring."""
    def _col(name):
        if hasattr(same_type_records, "columns"):
            return [_scr_num(v) for v in same_type_records.get(name, [])]
        return [_scr_num(r.get(name)) for r in same_type_records]
    return {
        "review_counts": [v for v in _col("review_count") if v is not None],
        "sales": [v for v in _col("monthly_sales") if v is not None],
    }


# ==============================================================================
# js_reader
# ==============================================================================
# Maps internal (canonical) names to the JS export headers, with broad alias
# fallback for minor version/locale differences.
_JSR_COLUMN_MAPPINGS: Dict[str, List[str]] = {
    "asin": ["ASIN", "Asin", "Product ASIN"],
    "title": ["Product Name", "Title", "Product Title", "Name"],
    "brand": ["Brand", "Brand Name"],
    "price": ["Price", "Current Price", "Buy Box Price"],
    "monthly_sales": [
        "Monthly Units Sold", "Est. Monthly Sales", "Monthly Sales",
        "Estimated Monthly Sales", "Units Sold",
    ],
    "daily_sales": ["Daily Units Sold", "Est. Daily Sales", "Daily Sales"],
    "monthly_revenue": [
        "Monthly Revenue", "Est. Monthly Revenue", "Revenue", "Estimated Revenue",
    ],
    "net_revenue": ["Net Revenue", "Net Sales"],
    "date_first_available": [
        "Date First Available", "First Available", "Listing Created",
        "Date Available",
    ],
    "rating": ["Star Rating", "Rating", "Average Rating", "Stars"],
    "reviews": ["Reviews", "Review Count", "Number of Reviews", "# of Reviews"],
    "amazon_fees": ["Amazon Fees", "FBA Fees", "Fees", "Estimated Fees"],
    "bsr": ["BSR", "Best Seller Rank", "Sales Rank", "Rank"],
    "lqs": ["LQS", "Listing Quality Score"],
    "fulfillment": ["Fulfillment", "Fulfilment", "Fulfillment Type", "Seller Type"],
    "num_sellers": ["No. of Sellers", "Number of Sellers", "Sellers", "# of Sellers"],
    "category": ["Category", "Product Category", "Parent Category"],
    "product_tier": ["Product Tier", "Tier", "Size Tier"],
    "dimensions": ["Dimensions", "Product Dimensions", "Size"],
    "weight": ["Weight", "Item Weight", "Product Weight"],
    "link": ["Link", "URL", "Product URL", "Amazon URL"],
}

# Canonical columns parsed as numeric (strip currency/commas, JS "-" -> None).
_JSR_NUMERIC_COLUMNS = [
    "price", "monthly_sales", "daily_sales", "monthly_revenue", "net_revenue",
    "rating", "reviews", "amazon_fees", "bsr", "lqs", "num_sellers",
]



def _jsr_find_column(df: pd.DataFrame, canonical: str) -> Optional[str]:
    """Return the first df column that matches any known alias for *canonical*."""
    aliases = _JSR_COLUMN_MAPPINGS.get(canonical, [])
    df_lower: Dict[str, str] = {str(c).lower().strip(): c for c in df.columns}

    for alias in aliases:
        if alias in df.columns:
            return alias
        if alias.lower().strip() in df_lower:
            return df_lower[alias.lower().strip()]

    # Broad partial-match fallback, but guard against over-eager substring hits
    # (e.g. "Net Revenue" should not be picked up for "monthly_revenue").
    for col in df.columns:
        col_strip = str(col).lower().strip()
        for alias in aliases:
            if col_strip == alias.lower() or col_strip.startswith(alias.lower()):
                return col
    return None


def _jsr_parse_numeric(value) -> Optional[float]:
    """Convert a cell value (possibly a formatted string or JS dash) to float."""
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ("-", "–", "—", "N/A", "n/a", ""):
        return None
    s = (
        s.replace(",", "").replace("$", "").replace("£", "").replace("€", "")
        .replace("%", "").replace("�", "")
    )
    # "1.2 kg", "14.3 x 31 cm" etc. — take the leading number if present.
    m = re.match(r"^-?\d+(\.\d+)?", s)
    if m:
        try:
            return float(m.group(0))
        except (ValueError, TypeError):
            return None
    return None


def _jsr_clean_asin(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper()
    return s if _ASIN_RE.match(s) else None


def read_js_products(file_path: str) -> Tuple[pd.DataFrame, int]:
    """
    Read a Jungle Scout keyword-search CSV/XLSX export and return a normalised
    DataFrame.

    Returns
    -------
    df : pd.DataFrame
        Normalised product data with canonical column names. Rows without a
        valid ASIN are dropped. Unmapped columns are preserved with a 'raw_'
        prefix (never discarded).
    row_count : int
        Number of product rows read.
    """
    path = Path(file_path.strip().strip('"').strip("'"))
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        df_raw = pd.read_csv(path)
    elif suffix in (".xlsx", ".xls"):
        df_raw = pd.read_excel(path, engine="openpyxl")
    else:
        raise ValueError(f"Unsupported format '{suffix}'. Supply .csv, .xlsx, or .xls")

    cols_used: Dict[str, str] = {}
    data: Dict[str, pd.Series] = {}
    for canonical in _JSR_COLUMN_MAPPINGS:
        actual = _jsr_find_column(df_raw, canonical)
        if actual and actual not in cols_used.values():
            cols_used[canonical] = actual
            data[canonical] = df_raw[actual].reset_index(drop=True)

    df = pd.DataFrame(data)

    # Append unmapped raw columns for reference (prefixed with 'raw_')
    mapped_actuals = set(cols_used.values())
    for col in df_raw.columns:
        if col not in mapped_actuals:
            safe_name = f"raw_{re.sub(r'[^a-zA-Z0-9_]', '_', str(col))}"
            df[safe_name] = df_raw[col].reset_index(drop=True)

    for col in _JSR_NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = df[col].apply(_jsr_parse_numeric)

    # Keep only rows with a valid ASIN.
    if "asin" in df.columns:
        df["asin"] = df["asin"].apply(_jsr_clean_asin)
        df = df[df["asin"].notna()].reset_index(drop=True)

    row_count = len(df)

    print(f"\n  JS products loaded : {row_count}")
    detected = [k for k in _JSR_COLUMN_MAPPINGS if k in cols_used]
    print(f"  Columns mapped     : {', '.join(detected) or '(none — check headers)'}")
    if "asin" not in cols_used:
        print("  WARNING: no ASIN column detected — verify this is a JS product export.")
    if "monthly_sales" not in cols_used:
        print("  WARNING: no monthly-sales column detected — sales ranking will be limited.")

    return df, row_count


# ==============================================================================
# xray_product_reader
# ==============================================================================
_XPR_COLUMN_MAPPINGS: Dict[str, List[str]] = {
    "asin": ["ASIN", "Asin"],
    "title": ["Product Details", "Title", "Product Name", "Product Title"],
    "url": ["URL", "Product URL", "Link"],
    "image_url": ["Image URL", "Image Url", "Main Image"],
    "brand": ["Brand", "Brand Name"],
    "price": ["Price"],
    "parent_sales": ["Parent Level Sales", "Parent Sales"],
    "asin_sales": ["ASIN Sales", "Asin Sales"],
    "recent_purchases": ["Recent Purchases", "Bought in past month"],
    "parent_revenue": ["Parent Level Revenue", "Parent Revenue"],
    "asin_revenue": ["ASIN Revenue", "Asin Revenue"],
    "title_char_count": ["Title Char. Count", "Title Character Count", "Title Length"],
    "bsr": ["BSR", "Best Seller Rank", "Sales Rank"],
    "seller_country": ["Seller Country/Region", "Seller Country", "Country"],
    "fees": ["Fees"],
    "active_sellers": ["Active Sellers", "Sellers", "No. of Sellers"],
    "rating": ["Ratings", "Rating", "Star Rating", "Stars"],
    "review_count": ["Review Count", "Reviews", "# of Reviews"],
    "images": ["Images", "Image Count", "# of Images"],
    "review_velocity": ["Review velocity", "Review Velocity", "Reviews/Month"],
    "buy_box": ["Buy Box", "Buy Box Seller", "BuyBox"],
    "category": ["Category", "Product Category"],
    "size_tier": ["Size Tier", "Product Tier", "Tier"],
    "fulfillment": ["Fulfillment", "Fulfilment", "Fulfillment Type"],
    "dimensions": ["Dimensions", "Product Dimensions"],
    "weight": ["Weight", "Item Weight"],
    "aba_most_clicked": ["ABA Most Clicked", "ABA Click Share"],
    "creation_date": ["Creation Date", "Date First Available", "Listed Since"],
    "sponsored": ["Sponsored", "Sponsored Type", "Ad Type"],
    "best_seller": ["Best Seller", "Bestseller", "Best Seller Badge"],
    "seller_age_mo": ["Seller Age (mo)", "Seller Age", "Seller Age (months)"],
    "seller": ["Seller", "Seller Name"],
}

# Parsed as numeric (strip currency/commas/glyph, H10 "-" -> None).
_XPR_NUMERIC_COLUMNS = [
    "price", "parent_sales", "asin_sales", "recent_purchases", "parent_revenue",
    "asin_revenue", "title_char_count", "bsr", "fees", "active_sellers", "rating",
    "review_count", "images", "review_velocity", "weight", "seller_age_mo",
]



def _xpr_find_column(df: pd.DataFrame, canonical: str) -> Optional[str]:
    aliases = _XPR_COLUMN_MAPPINGS.get(canonical, [])
    df_lower: Dict[str, str] = {str(c).lower().strip(): c for c in df.columns}
    for alias in aliases:
        if alias in df.columns:
            return alias
        if alias.lower().strip() in df_lower:
            return df_lower[alias.lower().strip()]
    # Prefix match handles trailing glyphs/units (e.g. "Price  �").
    for col in df.columns:
        col_strip = str(col).lower().strip()
        for alias in aliases:
            if col_strip == alias.lower() or col_strip.startswith(alias.lower()):
                return col
    return None


def _xpr_parse_numeric(value) -> Optional[float]:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if s in ("-", "–", "—", "N/A", "n/a", ""):
        return None
    s = (
        s.replace(",", "").replace("$", "").replace("£", "").replace("€", "")
        .replace("%", "").replace("�", "").strip()
    )
    m = re.match(r"^-?\d+(\.\d+)?", s)
    if m:
        try:
            return float(m.group(0))
        except (ValueError, TypeError):
            return None
    return None


def _xpr_clean_asin(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip().upper()
    return s if _ASIN_RE.match(s) else None


def classify_sponsored(value) -> Optional[str]:
    """Map an Xray `Sponsored` cell to a canonical ad class.

    Returns 'sponsored_brand' for Sponsored Brand / Sponsored Brand Video
    (the ex-"AMS" banner ads), 'sponsored_product' for plain "Sponsored", or
    None for organic / blank.
    """
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    s = str(value).strip().lower()
    if not s or s in ("-", "n/a", "none", "organic"):
        return None
    if s.startswith("sponsored brand"):
        return "sponsored_brand"
    if s.startswith("sponsored"):
        return "sponsored_product"
    return None


def read_xray_products(file_path: str) -> Tuple[pd.DataFrame, int]:
    """
    Read one H10 Xray Products XLSX export and return a normalised DataFrame.

    Adds a canonical `sponsored_class` column ('sponsored_brand' /
    'sponsored_product' / None) derived from the raw `Sponsored` cell. Rows
    without a valid ASIN are dropped. Unmapped columns are kept with a 'raw_'
    prefix.
    """
    path = Path(file_path.strip().strip('"').strip("'"))
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        df_raw = pd.read_excel(path)  # pandas handles the embedded drawings
    elif suffix == ".csv":
        df_raw = pd.read_csv(path)
    else:
        raise ValueError(f"Unsupported format '{suffix}'. Supply .xlsx, .xls, or .csv")

    cols_used: Dict[str, str] = {}
    data: Dict[str, pd.Series] = {}
    for canonical in _XPR_COLUMN_MAPPINGS:
        actual = _xpr_find_column(df_raw, canonical)
        if actual and actual not in cols_used.values():
            cols_used[canonical] = actual
            data[canonical] = df_raw[actual].reset_index(drop=True)

    df = pd.DataFrame(data)

    mapped_actuals = set(cols_used.values())
    for col in df_raw.columns:
        if col not in mapped_actuals:
            safe_name = f"raw_{re.sub(r'[^a-zA-Z0-9_]', '_', str(col))}"
            df[safe_name] = df_raw[col].reset_index(drop=True)

    for col in _XPR_NUMERIC_COLUMNS:
        if col in df.columns:
            df[col] = df[col].apply(_xpr_parse_numeric)

    if "sponsored" in df.columns:
        df["sponsored_class"] = df["sponsored"].apply(classify_sponsored)
    else:
        df["sponsored_class"] = None

    if "asin" in df.columns:
        df["asin"] = df["asin"].apply(_xpr_clean_asin)
        df = df[df["asin"].notna()].reset_index(drop=True)

    row_count = len(df)

    print(f"\n  Xray products loaded : {row_count}")
    detected = [k for k in _XPR_COLUMN_MAPPINGS if k in cols_used]
    print(f"  Columns mapped       : {', '.join(detected) or '(none — check headers)'}")
    if "sponsored" not in cols_used:
        print("  WARNING: no 'Sponsored' column detected — Sponsored Products / "
              "Sponsored Brands scoring will be unavailable from this file.")
    return df, row_count


def aggregate_sponsored(per_keyword: Dict[str, pd.DataFrame]) -> Dict[str, Dict]:
    """
    Aggregate sponsored presence per ASIN across multiple keyword exports.

    *per_keyword* maps a keyword label -> its read_xray_products() DataFrame.
    Returns ``{asin: {"sp_keywords": [...], "sb_keywords": [...],
    "sp_count": int, "sb_count": int, "total_keywords": int}}`` where sp = a
    Sponsored Product placement and sb = a Sponsored Brand placement, listing
    the keywords on which the ASIN advertised. This feeds the "ads on many
    high-volume keywords" rubric criteria.
    """
    agg: Dict[str, Dict] = {}
    total_keywords = len(per_keyword)
    for kw, df in per_keyword.items():
        if "asin" not in df.columns:
            continue
        for _, row in df.iterrows():
            asin = row.get("asin")
            if not asin:
                continue
            rec = agg.setdefault(
                asin, {"sp_keywords": [], "sb_keywords": []}
            )
            cls = row.get("sponsored_class")
            if cls == "sponsored_product" and kw not in rec["sp_keywords"]:
                rec["sp_keywords"].append(kw)
            elif cls == "sponsored_brand" and kw not in rec["sb_keywords"]:
                rec["sb_keywords"].append(kw)

    for asin, rec in agg.items():
        rec["sp_count"] = len(rec["sp_keywords"])
        rec["sb_count"] = len(rec["sb_keywords"])
        rec["total_keywords"] = total_keywords
    return agg


# ==============================================================================
# claude_client
# ==============================================================================
MODEL = "claude-opus-4-8"

# Reuse Phase 2's profile builder + schema (same relevancy-anchor concept).

# Max points per judgement-scored rubric element (the 184-point rubric).
LISTING_ELEMENT_MAX = {
    "product_images": 20,
    "product_title": 5,
    "bullets_description": 5,
    "unique_design": 2,
    "marketing_images": 15,
    "pricing_strategy": 10,
}


class Phase3ApiError(RuntimeError):
    """Raised when the Claude API cannot complete a Phase 3 judgement."""


# ──────────────────────────────────────────────────────────────────────────────
# Client construction (.env key)
# ──────────────────────────────────────────────────────────────────────────────

def load_client():
    """Build an Anthropic client from the project-root `.env`. No fallback."""
    try:
        from dotenv import load_dotenv
    except ImportError as e:
        raise Phase3ApiError("python-dotenv is not installed. Run setup again.") from e
    try:
        import anthropic
    except ImportError as e:
        raise Phase3ApiError("The 'anthropic' package is not installed. Run setup again.") from e

    load_dotenv(dotenv_path=_ENV_PATH)
    import os
    if not os.environ.get("ANTHROPIC_API_KEY"):
        raise Phase3ApiError(
            f"ANTHROPIC_API_KEY not found. Add it to {_ENV_PATH} (git-ignored)."
        )
    try:
        return anthropic.Anthropic()
    except Exception as e:  # noqa: BLE001
        raise Phase3ApiError(f"Could not initialise the Anthropic client: {e}") from e


# ──────────────────────────────────────────────────────────────────────────────
# Call 1 — same-product-type filter (Step 8)
# ──────────────────────────────────────────────────────────────────────────────

class SameTypeResult(BaseModel):
    same_type_asins: List[str] = Field(
        description="ASINs from the candidate list that are the SAME core product type as the target. Copy ASINs verbatim."
    )
    note: str = Field(description="One or two sentences on what was kept vs excluded and why.")


_SAME_TYPE_RULES = (
    "You are an Amazon competitor analyst. You are given a TARGET product profile "
    "and a list of candidate products (ASIN + title) returned by keyword searches.\n\n"
    "Decide which candidates are the SAME CORE PRODUCT TYPE as the target — a real "
    "competitor a shopper would cross-shop against this exact product. Use the "
    "target profile's type, key attributes and 'NOT this' list.\n\n"
    "EXCLUDE: accessories, spare parts, look-alikes of a different type, bundles of "
    "a different product, and items from the profile's 'NOT this' list, even if they "
    "appear in the same search results. Keep both FBA and FBM sellers — fulfilment "
    "is irrelevant to product type.\n\n"
    "Return only the ASINs that ARE the same product type. Copy ASINs verbatim; never "
    "invent one. If unsure about a candidate, exclude it and say so in the note."
)


def filter_same_product_type(
    client,
    profile: ProductProfile,
    candidates: pd.DataFrame,
    max_candidates: int = 250,
) -> SameTypeResult:
    """Return the subset of *candidates* that are the same product type as the
    target. *candidates* needs 'asin' and 'title' columns."""
    if candidates.empty:
        raise Phase3ApiError("No candidates to classify for same-product-type.")

    rows = candidates.head(max_candidates)
    listing = "\n".join(
        f"  - {r['asin']}: {str(r.get('title') or '')[:140]}" for _, r in rows.iterrows()
    )
    profile_block = (
        f"TARGET PRODUCT PROFILE:\n  Name: {profile.product_name}\n"
        f"  Type: {profile.product_type}\n"
        f"  Key attributes: {', '.join(profile.key_attributes) or '(none)'}\n"
        f"  NOT this: {', '.join(profile.not_this) or '(none)'}"
    )
    system = [
        {"type": "text", "text": _SAME_TYPE_RULES, "cache_control": {"type": "ephemeral"}},
        {"type": "text", "text": profile_block, "cache_control": {"type": "ephemeral"}},
    ]
    user = f"CANDIDATES (ASIN: title):\n{listing}\n\nReturn the ASINs that are the same product type."
    try:
        resp = client.messages.parse(
            model=MODEL, max_tokens=16000, thinking={"type": "adaptive"},
            output_config={"effort": "medium"}, system=system,
            messages=[{"role": "user", "content": user}], output_format=SameTypeResult,
        )
    except Exception as e:  # noqa: BLE001
        raise Phase3ApiError(f"Claude same-product-type filter failed: {e}") from e
    if resp.parsed_output is None:
        raise Phase3ApiError("Claude returned no parseable same-product-type result.")
    return resp.parsed_output


# ──────────────────────────────────────────────────────────────────────────────
# Call 2 — listing quality scoring (vision)
# ──────────────────────────────────────────────────────────────────────────────

class ListingScores(BaseModel):
    product_images: int = Field(description="Product-page image quality score 0-20: professional/lifestyle/infographics/count.")
    product_images_reason: str
    product_title: int = Field(description="Title quality score 0-5: relevant, clear, well-structured (NOT keyword-stuffed).")
    product_title_reason: str
    bullets_description: int = Field(description="Bullets/description quality score 0-5: benefits, keywords, addresses pain points.")
    bullets_description_reason: str
    unique_design: int = Field(description="Design uniqueness score 0-2: 2 = visibly unique/differentiated design, 0 = generic/saturated. Ignore patents.")
    unique_design_reason: str
    marketing_images: int = Field(description="Overall brand/marketing imagery score 0-15 from the listing gallery + A+ visible (storefront not available).")
    marketing_images_reason: str
    pricing_strategy: int = Field(description="Pricing-strategy strength 0-10 from price positioning vs the market; use -1 if it cannot be judged without the seller's full catalog.")
    pricing_strategy_reason: str


_LISTING_RULES = (
    "You are scoring ONE Amazon competitor's listing for a competitor-analysis "
    "rubric. Higher score = STRONGER competitor (harder for a new seller to beat).\n\n"
    "Score each element within its max. Judge images from the provided listing "
    "images (form, professionalism, lifestyle shots, infographics, count). Judge "
    "title and bullets from the text. For 'unique_design', 2 = a visibly "
    "differentiated/distinctive design, 0 = a generic design sold by many; ignore "
    "patents (the user verifies those separately). For 'marketing_images', judge "
    "the overall brand imagery quality from the gallery and any A+ visible — note "
    "you do NOT have the brand storefront. For 'pricing_strategy', judge from the "
    "price positioning given; if you genuinely cannot assess it without the "
    "seller's full catalogue, return -1 (do not guess).\n\n"
    "Give a one-sentence reason per element. Base every judgement only on the "
    "evidence provided — never invent details you cannot see."
)


def score_listing(
    client,
    competitor: Dict,
    image_urls: Optional[List[str]] = None,
    max_images: int = 5,
) -> ListingScores:
    """Score the judgement-based rubric elements for one competitor.

    *competitor* should carry: title, bullets (list), price, market_price_context
    (a short string e.g. 'cheapest of 3' or price list), a_plus, has_video,
    images_count. *image_urls* are the listing image URLs (sent to vision).
    """
    bullets = competitor.get("bullets") or []
    text = (
        f"COMPETITOR LISTING\n"
        f"Title: {competitor.get('title') or '(none)'}\n"
        f"Price: {competitor.get('price')}\n"
        f"Price context vs market: {competitor.get('market_price_context') or '(n/a)'}\n"
        f"A+ content present: {competitor.get('a_plus')}\n"
        f"Video present: {competitor.get('has_video')}\n"
        f"Image count: {competitor.get('images_count')}\n"
        f"Bullets:\n" + "\n".join(f"  - {b}" for b in bullets[:8])
    )
    content: List[Dict] = [{"type": "text", "text": text}]
    for u in (image_urls or [])[:max_images]:
        content.append({"type": "image", "source": {"type": "url", "url": u}})

    try:
        resp = client.messages.parse(
            model=MODEL, max_tokens=3000, thinking={"type": "adaptive"},
            output_config={"effort": "medium"},
            system=[{"type": "text", "text": _LISTING_RULES, "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": content}],
            output_format=ListingScores,
        )
    except Exception as e:  # noqa: BLE001
        raise Phase3ApiError(f"Claude listing-quality scoring failed: {e}") from e
    if resp.parsed_output is None:
        raise Phase3ApiError("Claude returned no parseable listing scores.")

    # Clamp each score to its element max (defensive; -1 sentinel kept as Unknown).
    s = resp.parsed_output
    for key, mx in LISTING_ELEMENT_MAX.items():
        val = getattr(s, key)
        if val is not None and val >= 0:
            setattr(s, key, min(val, mx))
    return s


# ──────────────────────────────────────────────────────────────────────────────
# Call 3 — review analysis (weaknesses / solutions / strengths)
# ──────────────────────────────────────────────────────────────────────────────

class ReviewAnalysis(BaseModel):
    weaknesses: List[str] = Field(description="Top 3 recurring complaints/pain points from 1-3 star reviews. Empty if no review data.")
    solutions: List[str] = Field(description="A concrete product solution for each weakness (materials/specs/design), aligned by index.")
    strengths: List[str] = Field(description="Top 3 features/materials/design aspects customers praise in 4-5 star reviews. Empty if no review data.")
    note: str = Field(description="Caveats; if no review data was provided, say 'Unknown — need user confirmation'.")


_REVIEW_RULES = (
    "You analyse Amazon customer feedback for a competitor. The input has two "
    "labelled sections:\n"
    "  - 'STRENGTHS (Amazon AI insight)': Amazon's 'Customers say' AI summary of "
    "what customers like. Derive the top 3 praised strengths from THIS section.\n"
    "  - 'WEAKNESSES (1-2 star reviews)': the actual 1-2 star reviews. Extract the "
    "top 3 recurring complaints from THESE; for each, propose a concrete product "
    "solution (material, specification, or design change).\n\n"
    "Base everything strictly on the supplied text. If a section is missing, leave "
    "its list empty and note it as 'Unknown — need user confirmation' — never "
    "invent complaints or praise (Rule 0)."
)


def analyze_reviews(client, competitor_label: str, reviews_text: str = "") -> ReviewAnalysis:
    """Analyse a competitor's reviews. With no review text, returns an explicit
    Unknown result (no fabrication)."""
    if not (reviews_text or "").strip():
        return ReviewAnalysis(
            weaknesses=[], solutions=[], strengths=[],
            note="Unknown — need user confirmation (no review data supplied).",
        )
    user = f"Competitor: {competitor_label}\n\nREVIEWS:\n{reviews_text[:15000]}"
    try:
        resp = client.messages.parse(
            model=MODEL, max_tokens=2500, thinking={"type": "adaptive"},
            output_config={"effort": "medium"},
            system=[{"type": "text", "text": _REVIEW_RULES, "cache_control": {"type": "ephemeral"}}],
            messages=[{"role": "user", "content": user}], output_format=ReviewAnalysis,
        )
    except Exception as e:  # noqa: BLE001
        raise Phase3ApiError(f"Claude review analysis failed: {e}") from e
    if resp.parsed_output is None:
        raise Phase3ApiError("Claude returned no parseable review analysis.")
    return resp.parsed_output


# ==============================================================================
# competitor_selector
# ==============================================================================
def _csel_num(v) -> Optional[float]:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _norm_fulfillment(raw) -> str:
    """Normalise fulfilment codes to FBA / FBM / AMZ across JS and Xray."""
    s = str(raw or "").strip().upper()
    if s in ("FBA",):
        return "FBA"
    if s in ("FBM", "MFN"):
        return "FBM"
    if s in ("AMZ", "AMAZON"):
        return "AMZ"
    return s or "Unknown"


def build_candidate_records(
    js_df: Optional[pd.DataFrame],
    per_keyword_xray: Dict[str, pd.DataFrame],
    sponsored_agg: Dict[str, Dict],
    pool: List[str],
) -> pd.DataFrame:
    """One merged record per ASIN in *pool* (JS + Xray fields + sponsored agg).

    Sales source priority: JS monthly_sales (the workflow's estimate), else the
    Xray ASIN-level sales. For each ASIN the representative Xray row is the one
    with the highest ASIN sales across the keyword exports.
    """
    js_by_asin: Dict[str, pd.Series] = {}
    if js_df is not None and "asin" in js_df.columns:
        for _, r in js_df.iterrows():
            js_by_asin.setdefault(r["asin"], r)

    # Best Xray row per ASIN + presence across keyword files. Presence counts
    # DISTINCT keywords (an ASIN can appear as both an organic and a sponsored
    # row within one export, which must not double-count).
    xray_best: Dict[str, pd.Series] = {}
    presence_kw: Dict[str, set] = {}
    for kw, df in per_keyword_xray.items():
        if "asin" not in df.columns:
            continue
        for _, r in df.iterrows():
            a = r["asin"]
            presence_kw.setdefault(a, set()).add(kw)
            cur = xray_best.get(a)
            if cur is None or (_csel_num(r.get("asin_sales")) or 0) > (_csel_num(cur.get("asin_sales")) or 0):
                xray_best[a] = r
    presence = {a: len(kws) for a, kws in presence_kw.items()}

    records = []
    for asin in pool:
        js = js_by_asin.get(asin)
        xr = xray_best.get(asin)
        spons = sponsored_agg.get(asin, {})

        def pick(js_key, xr_key=None, prefer="js"):
            xr_key = xr_key or js_key
            jv = js.get(js_key) if js is not None else None
            xv = xr.get(xr_key) if xr is not None else None
            order = (jv, xv) if prefer == "js" else (xv, jv)
            for v in order:
                if v is not None and not (isinstance(v, float) and np.isnan(v)):
                    return v
            return None

        monthly_sales = _csel_num(pick("monthly_sales", "asin_sales"))
        monthly_revenue = _csel_num(pick("monthly_revenue", "asin_revenue"))
        fulfillment = _norm_fulfillment(pick("fulfillment", "fulfillment"))

        records.append({
            "asin": asin,
            "title": pick("title", "title"),
            "brand": pick("brand", "brand"),
            "price": _csel_num(pick("price", "price")),
            "monthly_sales": monthly_sales,
            "monthly_revenue": monthly_revenue,
            "rating": _csel_num(pick("rating", "rating", prefer="xr")),
            "review_count": _csel_num(pick("reviews", "review_count", prefer="xr")),
            "review_velocity": _csel_num(xr.get("review_velocity")) if xr is not None else None,
            "best_seller": (str(xr.get("best_seller")).strip() if xr is not None and xr.get("best_seller") is not None else None),
            "bsr": _csel_num(pick("bsr", "bsr", prefer="xr")),
            "date_first_available": pick("date_first_available", "creation_date"),
            "fulfillment": fulfillment,
            "seller": xr.get("seller") if xr is not None else None,
            "seller_age_mo": _csel_num(xr.get("seller_age_mo")) if xr is not None else None,
            "active_sellers": _csel_num(pick("num_sellers", "active_sellers", prefer="xr")),
            "image_url": (xr.get("image_url") if xr is not None else None),
            "url": pick("link", "url", prefer="xr"),
            "dimensions": pick("dimensions", "dimensions"),
            "weight": pick("weight", "weight"),
            "category": pick("category", "category"),
            "sp_count": spons.get("sp_count", 0),
            "sb_count": spons.get("sb_count", 0),
            "sp_keywords": spons.get("sp_keywords", []),
            "sb_keywords": spons.get("sb_keywords", []),
            "keyword_presence": presence.get(asin, 0),
            "total_keywords": len(per_keyword_xray),
            "in_js": js is not None,
            "in_xray": xr is not None,
        })

    return pd.DataFrame(records)


def select_top_competitors(
    client,
    profile: ProductProfile,
    records: pd.DataFrame,
    n: int = 3,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    Return (top_n, all_same_type, report_lines).

    Claude filters to same-product-type; Python ranks the survivors by monthly
    sales (tie-break: keyword presence, then review count) and takes the top *n*.
    Both FBA and FBM are eligible — fulfilment is never used to exclude.
    """
    report: List[str] = []
    if records.empty:
        raise ValueError("No candidate records to select from.")

    result = filter_same_product_type(client, profile, records[["asin", "title"]])
    kept = set(result.same_type_asins)
    same = records[records["asin"].isin(kept)].copy()
    report.append(f"  Candidates considered : {len(records)}")
    report.append(f"  Same product type     : {len(same)}  (Claude)")
    report.append(f"  Note                  : {result.note}")

    if same.empty:
        return same, same, report

    same["_sales"] = same["monthly_sales"].fillna(0)
    same["_pres"] = same["keyword_presence"].fillna(0)
    same["_rev"] = same["review_count"].fillna(0)
    same = same.sort_values(["_sales", "_pres", "_rev"], ascending=False).reset_index(drop=True)
    same = same.drop(columns=["_sales", "_pres", "_rev"])

    # One competitor per brand: `same` is sorted best-first, so keep only the
    # first (strongest) listing of each brand and drop the rest — that way the
    # top 3 are distinct brands, not multiple variants of the same brand's
    # product. Rows with no brand are all kept.
    deduped = same
    if "brand" in same.columns:
        seen_brands: set = set()
        keep: List = []
        for idx, brand in same["brand"].items():
            b = str(brand or "").strip().lower()
            if b and b in seen_brands:
                continue
            if b:
                seen_brands.add(b)
            keep.append(idx)
        deduped = same.loc[keep]
        dropped = len(same) - len(deduped)
        if dropped:
            report.append(f"  Same-brand variants dropped : {dropped} (kept best per brand)")

    top = deduped.head(n).copy()
    ff = ", ".join(f"{r.asin}={r.fulfillment} ({r.brand})" for r in top.itertuples())
    report.append(f"  Top {len(top)} by sales      : {ff}")
    return top, same, report


# ==============================================================================
# reviews_flyby
# ==============================================================================
HOST = "real-time-amazon-data.p.rapidapi.com"
MAX_STRENGTHS = 5
MAX_PAGES_PER_STAR = 10  # 10 reviews/page -> up to ~100 reviews per star band

# Amazon marketplace suffix -> API country code.
_COUNTRY = {"co.uk": "GB", "com": "US", "de": "DE", "fr": "FR", "it": "IT",
            "es": "ES", "ca": "CA", "com.au": "AU", "co.jp": "JP", "in": "IN"}


class FlybyError(RuntimeError):
    """Raised when the reviews API cannot complete a request."""


def _env(name: str) -> str:
    try:
        from dotenv import load_dotenv
        load_dotenv(dotenv_path=_ENV_PATH)
    except Exception:  # noqa: BLE001
        pass
    return (os.environ.get(name) or "").strip()


def _amazon_cookie() -> str:
    """The Amazon session cookie for the reviews API, from .env.

    Preferred form: the split variables ``AMAZON_SESSION_ID`` /
    ``AMAZON_UBID_ACBUK`` / ``AMAZON_X_ACBUK`` / ``AMAZON_AT_ACBUK`` (values
    copied one-by-one from the browser's cookie panel for amazon.co.uk),
    joined into the header string here. They take priority over a full
    ``AMAZON_COOKIE`` header string so a stale machine-level AMAZON_COOKIE
    can't shadow freshly updated .env values.
    """
    parts = (
        ("session-id", _env("AMAZON_SESSION_ID")),
        ("ubid-acbuk", _env("AMAZON_UBID_ACBUK")),
        ("x-acbuk", _env("AMAZON_X_ACBUK")),
        ("at-acbuk", _env("AMAZON_AT_ACBUK")),
    )
    split = "; ".join(f"{n}={v}" for n, v in parts if v)
    return split or _env("AMAZON_COOKIE")


def _rvf_get(path: str, key: str, params: Dict, timeout: float) -> Dict:
    r = requests.get(
        f"https://{HOST}{path}",
        headers={"X-RapidAPI-Key": key, "X-RapidAPI-Host": HOST},
        params=params, timeout=timeout,
    )
    if r.status_code != 200:
        raise FlybyError(f"{path} HTTP {r.status_code}: {r.text[:200]}")
    return (r.json() or {}).get("data") or {}


def _rvf_norm(rev: Dict) -> Tuple[str, Dict]:
    """Normalise one API review into (id, {stars,title,body,verified,date})."""
    rid = rev.get("review_id") or f"{rev.get('review_title')}|{str(rev.get('review_comment'))[:40]}"
    try:
        stars = int(round(float(rev.get("review_star_rating"))))
    except (TypeError, ValueError):
        stars = None
    return rid, {
        "stars": stars,
        "title": (rev.get("review_title") or "").strip(),
        "body": (rev.get("review_comment") or "").strip(),
        "verified": rev.get("is_verified_purchase"),
        "date": rev.get("review_date"),
    }


def fetch_reviews(asin: str, marketplace: str = "co.uk", timeout: float = 60) -> Dict:
    """Fetch + select reviews for one ASIN per the project rule
    (all 1-2★ / max five 4-5★ / no 3★).

    Returns {asin, marketplace, gated_access, total_ratings, distribution,
    strengths[...], weaknesses[...], weaknesses_count, note, source}.
    Never raises for normal failures — reports via the "error" key instead.
    """
    asin = (asin or "").strip().upper()
    if not re.fullmatch(r"[A-Z0-9]{10}", asin):
        return {"asin": asin, "error": f"Not a valid ASIN: {asin!r}",
                "weaknesses": [], "weaknesses_count": 0}
    key = _env("RAPIDAPI_KEY")
    if not key:
        return {"asin": asin, "error": f"RAPIDAPI_KEY not found. Add it to {_ENV_PATH}.",
                "weaknesses": [], "weaknesses_count": 0}
    cookie = _amazon_cookie()
    country = _COUNTRY.get(marketplace.lower(), marketplace.upper())

    dist: Dict[str, int] = {}
    total_ratings = None
    weaknesses: Dict[str, Dict] = {}
    strengths: Dict[str, Dict] = {}
    errors: List[str] = []
    gated = False
    cookie_dead = False

    def _page(params: Dict) -> List[Dict]:
        nonlocal dist, total_ratings
        data = _rvf_get("/product-reviews", key, params, timeout)
        dist = {str(k): v for k, v in (data.get("rating_distribution") or {}).items()} or dist
        total_ratings = data.get("total_ratings") or total_ratings
        return data.get("reviews") or []

    # ── 1. Weaknesses: ALL 1★ + 2★, paginated (needs the session cookie) ─────
    if cookie:
        for star_val, tag in (("1_STARS", 1), ("2_STARS", 2)):
            if cookie_dead:
                break
            for page in range(1, MAX_PAGES_PER_STAR + 1):
                try:
                    revs = _page({"asin": asin, "country": country, "star_rating": star_val,
                                  "sort_by": "MOST_RECENT", "page": page, "cookie": cookie})
                except Exception as e:  # noqa: BLE001
                    errors.append(str(e))
                    break
                if not revs:
                    break
                got = dict(_rvf_norm(r) for r in revs)
                # Filter honoured? Off-band stars on page 1 mean the cookie is
                # dead and Amazon served the unfiltered public sample.
                if any(v["stars"] not in (tag, None) for v in got.values()):
                    cookie_dead = True
                    break
                new = 0
                for rid, v in got.items():
                    if rid not in weaknesses and (v["title"] or v["body"]):
                        weaknesses[rid] = v
                        new += 1
                gated = True
                if new == 0:  # repeated page -> end of the list
                    break

    # ── 2. Strengths: one TOP_REVIEWS page, keep max five 4-5★ ───────────────
    try:
        revs = _page({"asin": asin, "country": country, "star_rating": "ALL",
                      "sort_by": "TOP_REVIEWS", "page": 1,
                      **({"cookie": cookie} if cookie and not cookie_dead else {})})
    except Exception as e:  # noqa: BLE001
        errors.append(str(e))
        revs = []
    sample_negatives = 0
    for r in revs:
        rid, v = _rvf_norm(r)
        if not (v["title"] or v["body"]):
            continue
        if v["stars"] in (4, 5) and len(strengths) < MAX_STRENGTHS and rid not in strengths:
            strengths[rid] = v
        # Cookieless fallback: salvage any 1-2★ present in the public sample.
        if v["stars"] in (1, 2) and rid not in weaknesses:
            weaknesses[rid] = v
            sample_negatives += 1
    # 3★ are excluded by design (user rule).

    if not weaknesses and not strengths and errors:
        msg = "; ".join(errors)
        if "not subscribed" in msg.lower():
            msg = ("RAPIDAPI_KEY is not subscribed to the Real-Time Amazon Data API "
                   "(rapidapi.com, API: real-time-amazon-data) - renew the subscription or "
                   f"update RAPIDAPI_KEY in {_ENV_PATH}. Reviews were skipped, so the "
                   "strengths/weaknesses columns fall back to 'Unknown'. Raw: " + msg)
        return {"asin": asin, "error": msg, "weaknesses": [], "weaknesses_count": 0}

    # ── 3. Honest note ────────────────────────────────────────────────────────
    note = ""
    try:
        neg_pct = int(dist.get("1") or 0) + int(dist.get("2") or 0)
    except (TypeError, ValueError):
        neg_pct = 0
    if cookie_dead:
        note = ("The Amazon session cookie appears expired/invalid — Amazon ignored the "
                "star filter, so only the public detail-page sample was available. Refresh "
                "the cookie values in .env (AMAZON_SESSION_ID / AMAZON_UBID_ACBUK / "
                "AMAZON_X_ACBUK / AMAZON_AT_ACBUK) for the full 1-2 star list.")
    elif not cookie:
        note = ("No Amazon session cookie in .env (AMAZON_SESSION_ID / AMAZON_UBID_ACBUK / "
                "AMAZON_X_ACBUK / AMAZON_AT_ACBUK, or a full AMAZON_COOKIE) — only the "
                "public detail-page sample was available.")
    if not gated and not weaknesses and neg_pct:
        note += (" " if note else "") + (
            f"{neg_pct}% of all ratings are 1-2 star but none were retrievable.")

    return {
        "asin": asin,
        "marketplace": marketplace,
        "gated_access": gated,
        "total_ratings": total_ratings,
        "distribution": dist,
        "strengths": list(strengths.values()),
        "weaknesses": list(weaknesses.values()),
        "weaknesses_count": len(weaknesses),
        "note": note,
        "source": ("flyby/rapidapi product-reviews — full gated 1-2★ list via session cookie"
                   if gated else
                   "flyby/rapidapi product-reviews — public detail-page sample only"),
    }


def reviews_text(result: Optional[Dict]) -> str:
    """Render a fetch result into the STRENGTHS/WEAKNESSES block consumed by
    `claude_client.analyze_reviews` (empty if nothing usable)."""
    if not result or result.get("error"):
        return ""
    out: List[str] = []
    dist = result.get("distribution") or {}
    if dist:
        out.append("=== RATING DISTRIBUTION (all ratings) ===")
        line = ", ".join(f"{s}*: {dist.get(s, 0)}%" for s in ("5", "4", "3", "2", "1"))
        if result.get("total_ratings"):
            line += f"  (of {result['total_ratings']} total ratings)"
        out.append(line)
    if result.get("strengths"):
        out.append("\n=== STRENGTHS (4-5 star reviews, max 5) ===")
        for i, r in enumerate(result["strengths"], start=1):
            head = (r.get("title", "") + " — ") if r.get("title") else ""
            out.append(f"Review {i}: [{r.get('stars')}*] {head}{r.get('body', '')}")
    if result.get("weaknesses"):
        src = "full 1-2 star list" if result.get("gated_access") else "detail-page sample"
        out.append(f"\n=== WEAKNESSES (1-2 star reviews — {src}) ===")
        for i, r in enumerate(result["weaknesses"], start=1):
            head = (r.get("title", "") + " — ") if r.get("title") else ""
            out.append(f"Review {i}: [{r.get('stars')}*] {head}{r.get('body', '')}")
    if result.get("note"):
        out.append(f"\nNote: {result['note']}")
    return "\n".join(out)


# ==============================================================================
# output
# ==============================================================================
# Column letters in the Competitor Analysis tab.
COL_GRADING, COL_INFO, COL_SCORE, COL_MAX = "E", "F", "J", "K"
COL_META_LABEL, COL_META_VALUE = "B", "C"

# Map a normalised column-E label -> scorer element key. Marketing-section
# duplicates (images/title/bullets) are disambiguated by section state.
_LABEL_KEYS = {
    "price": "price",
    "sponsored products": "sponsored_products",
    "ams ads": "sponsored_brands",
    "sponsored brands": "sponsored_brands",
    "sponsored brand": "sponsored_brands",
    "# of product reviews": "reviews",
    "avg. product rating": "rating",
    "age of product": "age",
    "bestseller badge": "bestseller",
    "amazon or 3p seller": "seller",
    "unique product design": "unique_design",
    "fba or fbm": "fba_fbm",
    "pricing strategy": "pricing_strategy",
    # new MARKET MOMENTUM rows:
    "review velocity": "review_velocity",
    "product variations": "variations",
    "sales / revenue strength": "sales_strength",
    "sales/revenue strength": "sales_strength",
    "enhanced content (a+/video)": "enhanced_content",
    "enhanced content": "enhanced_content",
}
# Section-sensitive labels: (product-page key, marketing key)
_SECTION_LABEL_KEYS = {
    "product images": ("product_images", "mkt_images"),
    "product title": ("product_title", "mkt_title"),
    "product description/bullet points": ("bullets_description", "mkt_bullets"),
}

_META_KEYS = {
    "brand name": "brand",
    "product page url": "url",
    "seller name": "seller_name",
    "units sales/month": "monthly_sales",
    "parent category": "category",
}


def _out_norm(v) -> str:
    return re.sub(r"\s+", " ", str(v or "").strip().lower()).rstrip(":")


def _find_block_starts(ws) -> List[int]:
    """Rows where each competitor block begins (col B '..Competitor')."""
    starts = []
    for r in range(1, ws.max_row + 1):
        b = ws.cell(row=r, column=2).value
        if b and re.search(r"\bcompetitor\b", str(b), re.I) and re.match(r"\s*\d", str(b)):
            starts.append(r)
    return starts


def _info_fallback(key: str, comp: Dict) -> Optional[str]:
    """Info-cell content for judgement elements whose scorer left no info text:
    the actual title for the Product Title rows, the media facts for the
    Product Images rows."""
    if key in ("product_title", "mkt_title"):
        return comp.get("title")
    if key in ("product_images", "mkt_images"):
        n = comp.get("images_count")
        if n is None:
            return None
        return f"{int(n)} listing image(s)" + (", video present" if comp.get("has_video") else "")
    if key in ("bullets_description", "mkt_bullets"):
        bullets = [str(b).strip() for b in (comp.get("bullets") or []) if str(b or "").strip()]
        if bullets:
            return "\n".join(f"- {b}" for b in bullets[:5])[:600]
        desc = (comp.get("description") or "").strip()
        return desc[:600] or None
    return None


def _embed_product_image(ws, row: int, comp: Dict) -> bool:
    """Embed the main product image in the Info cell of a Product Images
    grading row, scaled to the row height. Returns True on success."""
    urls = list(comp.get("image_urls") or [])
    if not urls and comp.get("image_url"):
        urls = [comp["image_url"]]
    if not urls:
        return False
    try:
        import io
        import requests
        from openpyxl.drawing.image import Image as XLImage
        resp = requests.get(urls[0], timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        resp.raise_for_status()
        img = XLImage(io.BytesIO(resp.content))
        row_pt = ws.row_dimensions[row].height or 60
        target = max(40.0, row_pt * 4 / 3 - 6)  # row height (pt) -> px, small margin
        scale = min(target / img.height, 1.0)
        img.width, img.height = int(img.width * scale), int(img.height * scale)
        ws.add_image(img, f"{COL_INFO}{row}")
        return True
    except Exception:  # noqa: BLE001 — caller falls back to text info
        return False


def _fill_product_box(ws, start: int, end: int, comp: Dict) -> None:
    """The tall empty merged box under the 'N Competitor' label (B11:C17 etc.)
    gets the PRODUCT TITLE (user decision — no image here)."""
    if not comp.get("title"):
        return
    for mr in sorted(ws.merged_cells.ranges, key=lambda m: m.min_row):
        if (mr.min_col == 2 and mr.max_col >= 3 and start < mr.min_row
                and mr.max_row < end and (mr.max_row - mr.min_row) >= 3):
            from openpyxl.styles import Alignment
            anchor = ws.cell(mr.min_row, mr.min_col)
            anchor.value = comp["title"]
            anchor.alignment = Alignment(wrap_text=True, vertical="center",
                                         horizontal="center")
            return


def _fill_block(ws, start: int, end: int, comp: Dict, scored: Dict, review) -> List[str]:
    """Fill one competitor block in [start, end). Returns notes about misses."""
    misses: List[str] = []
    elements = scored["elements"]
    in_marketing = False
    seen_keys = set()

    _fill_product_box(ws, start, end, comp)

    for r in range(start, end):
        # Section transition
        gtext = " ".join(
            _out_norm(ws.cell(row=r, column=c).value) for c in (5, 6, 7)
        )
        if "overall marketing strategy" in gtext:
            in_marketing = True

        # Metadata (col B label -> col C value)
        blabel = _out_norm(ws.cell(row=r, column=2).value)
        if blabel in _META_KEYS:
            key = _META_KEYS[blabel]
            if key == "seller_name":
                val = comp.get("seller") or comp.get("buy_box")
            elif key == "category":
                val = comp.get("subcategory") or comp.get("category")
            else:
                val = comp.get(key)
            if val is not None:
                ws[f"{COL_META_VALUE}{r}"] = val

        # Grading element (col E label)
        elabel = _out_norm(ws.cell(row=r, column=5).value)
        key = None
        if elabel in _SECTION_LABEL_KEYS:
            key = _SECTION_LABEL_KEYS[elabel][1 if in_marketing else 0]
        elif elabel in _LABEL_KEYS:
            key = _LABEL_KEYS[elabel]
        if key and key in elements and key not in seen_keys:
            seen_keys.add(key)
            el = elements[key]
            # Rescaled /100 weight -> col K, so SUM(K) (the Max Score) = 100.
            ws[f"{COL_MAX}{r}"] = el.get("max")
            if el["score"] is None:
                ws[f"{COL_INFO}{r}"] = el.get("info") or "Unknown — need user confirmation"
                ws[f"{COL_SCORE}{r}"] = None
            else:
                # The Price Info cell feeds numeric formulas in Pricing Analysis
                # (Average Retail Price -> FBA payout), so write the raw number.
                if key == "price" and comp.get("price") is not None:
                    info_cell = ws[f"{COL_INFO}{r}"]
                    info_cell.value = round(float(comp["price"]), 2)
                    info_cell.number_format = "£#,##0.00"
                elif key in ("product_images", "mkt_images") and _embed_product_image(ws, r, comp):
                    pass  # the image itself is the Info content
                elif el.get("info"):
                    ws[f"{COL_INFO}{r}"] = el["info"]
                else:
                    fb = _info_fallback(key, comp)
                    if fb:
                        ws[f"{COL_INFO}{r}"] = fb
                ws[f"{COL_SCORE}{r}"] = el["score"]

    # Weaknesses / solutions / strengths (label headers -> the "1./2./3." rows)
    _fill_review_section(ws, start, end, review)

    # Report any rubric element we never found a row for (e.g. new rows not added)
    missing_keys = [k for k in elements if k not in seen_keys]
    if missing_keys:
        misses.append(f"block@{start}: no row for {', '.join(missing_keys)}")
    return misses


def _is_review_header(e: str) -> Optional[str]:
    """Return 'weak' / 'sol' / 'strong' if E-label *e* is a review-section header."""
    if "solutions to top 3" in e:
        return "sol"  # must be checked first: contains "top 3 weakness" too
    if "top 3 weakness" in e:
        return "weak"
    if "top 3 strength" in e:
        return "strong"
    return None


def _fill_review_section(ws, start: int, end: int, review) -> None:
    """Write the 3 weaknesses / solutions / strengths under their headers.

    Each header (e.g. 'TOP 3 WEAKNESSES') is followed by '1.'/'2.'/'3.' rows in
    column E; we append the text to those. With no review data the first item
    carries the Unknown note (Rule 0).
    """
    lists = {
        "weak": list(getattr(review, "weaknesses", []) or []),
        "sol": list(getattr(review, "solutions", []) or []),
        "strong": list(getattr(review, "strengths", []) or []),
    }
    note = getattr(review, "note", "") or ""

    for r in range(start, end):
        kind = _is_review_header(_out_norm(ws.cell(row=r, column=5).value))
        if not kind:
            continue
        items = lists[kind]
        filled = 0
        rr = r + 1
        while rr < end and filled < 3:
            cell_e = str(ws.cell(row=rr, column=5).value or "").strip()
            if _is_review_header(_out_norm(cell_e)):
                break
            m = re.match(r"^([123])\.", cell_e)
            if m:
                idx = int(m.group(1)) - 1
                # Formatting only: make each item one single, middle-aligned cell
                # by merging E:I across the row (unmerge any overlap first).
                for mr in list(ws.merged_cells.ranges):
                    if mr.min_row <= rr <= mr.max_row and mr.min_col <= 9 and mr.max_col >= 5:
                        ws.unmerge_cells(str(mr))
                cell = ws.cell(row=rr, column=5)
                if idx < len(items):
                    cell.value = f"{idx + 1}. {items[idx]}"
                elif not items and idx == 0 and note:
                    cell.value = f"1. {note}"
                ws.merge_cells(start_row=rr, start_column=5, end_row=rr, end_column=9)
                from openpyxl.styles import Alignment
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                filled += 1
            rr += 1


def _fill_pricing_analysis(wb, comps: List[Dict], target_title: Optional[str] = None) -> None:
    """Clear Rule-0-violating placeholder costs and write the data we have.

    The competitor rows (13/14/15) pull Parent Category / URL / Brand / Price /
    Sales from the Competitor Analysis tab via existing formulas — those populate
    automatically. Here we (a) blank COGS/shipping/HTS placeholders (Rule 0 —
    blank until sourcing) and (b) write BSR, dimensions and weight from our data.
    """
    if "Pricing Analysis" not in wb.sheetnames:
        return
    ws = wb["Pricing Analysis"]

    # The TARGET product's title (from Phase 1) replaces the 'Electric Balloon
    # Pump' placeholder in the merged D11:F11 header cell.
    if target_title:
        target_cell = "D11"
        for row in ws.iter_rows(min_row=5, max_row=15):
            for c in row:
                if c.value and "balloon" in str(c.value).lower():
                    target_cell = c.coordinate
                    break
        ws[target_cell] = target_title

    # Repair the Top-3 row's template refs (#REF! in the original template) so
    # category/URL/brand/price/sales flow from the 3rd competitor block too.
    _BLOCK3_REFS = {
        "D15": "=+'Competitor Analysis'!C115",  # Parent Category
        "E15": "=+'Competitor Analysis'!C112",  # Product URL
        "F15": "=+'Competitor Analysis'!C111",  # Brand / Product Name
        "I15": "='Competitor Analysis'!F106",   # Avg Retail Price (price Info cell)
        "L15": "='Competitor Analysis'!C114",   # Projected Monthly Sales
    }
    for cell, formula in _BLOCK3_REFS.items():
        if isinstance(ws[cell].value, str) and "#REF" in ws[cell].value:
            ws[cell] = formula

    rows = [13, 14, 15]  # Top 1 / 2 / 3 competitor rows
    # Columns: G=Avg BSR, H=COGS, I=Avg Retail Price, J=FBA payout (sea),
    #          L=Projected sales, M/N=freight per unit, P=Shipping,
    #          X/Y=HTS, AB/AC/AD=Size cm, AG=Weight kg.
    for i, r in enumerate(rows):
        comp = comps[i] if i < len(comps) else None
        # Clear Rule-0 cost placeholders.
        for col in ("H", "P", "X", "Y"):
            ws[f"{col}{r}"] = None
        if comp is None:
            for col in ("G", "M", "N"):
                ws[f"{col}{r}"] = None
            continue
        # Average Parent BSR Rank = Keepa 90-day average (current BSR fallback).
        bsr = comp.get("bsr_90d")
        if not isinstance(bsr, (int, float)):
            bsr = comp.get("bsr")
        ws[f"G{r}"] = int(bsr) if isinstance(bsr, (int, float)) else None
        # FBA payout (header formula), with THIS competitor's actual Keepa fees
        # instead of the example product's hardcoded 15.3% / £2.87. The +0.5
        # supplier->warehouse shipping placeholder is kept from the template.
        pct, fee = comp.get("referral_pct"), comp.get("fba_pick_pack_fee")
        if isinstance(pct, (int, float)) and isinstance(fee, (int, float)):
            ws[f"J{r}"] = (f"=I{r}-((I{r}*{round(pct / 100.0, 4)})"
                           f"+(I{r}/6)+({round(fee, 2)})+(0.5))")
        dims = _parse_dims_cm(comp.get("dimensions"))
        if dims:
            ws[f"AB{r}"], ws[f"AC{r}"], ws[f"AD{r}"] = dims
        wt = _parse_weight_kg(comp.get("weight"), comp.get("package_weight_g"))
        if wt is not None:
            ws[f"AG{r}"] = wt
            # Header formula per unit: (sales x weight x 1.1 x rate)/sales -> the
            # sales terms cancel, leaving weight x 1.1 x $3 (sea) / $6 (air).
            ws[f"M{r}"] = f"=AG{r}*1.1*3"
            ws[f"N{r}"] = f"=AG{r}*1.1*6"
        else:
            ws[f"M{r}"], ws[f"N{r}"] = None, None  # clear the template's =26/#REF!


def _parse_dims_cm(value) -> Optional[List[float]]:
    if not value:
        return None
    nums = re.findall(r"\d+(?:\.\d+)?", str(value))
    if len(nums) >= 3:
        return [float(nums[0]), float(nums[1]), float(nums[2])]
    return None


def _parse_weight_kg(weight_str, package_weight_g) -> Optional[float]:
    if weight_str:
        s = str(weight_str).lower()
        m = re.search(r"(\d+(?:\.\d+)?)", s)
        if m:
            val = float(m.group(1))
            return val if "kg" in s else round(val / 1000.0, 3) if "g" in s else val
    if isinstance(package_weight_g, (int, float)) and package_weight_g > 0:
        return round(package_weight_g / 1000.0, 3)
    return None


def _fill_sponsored_products(wb, keywords: List[str]) -> None:
    """Write up to 6 launch keywords into the Sponsored Products tab, for each
    match type (Exact rows 7-12, Phrase 14-19, Broad 21-26). Bids stay blank."""
    if "Sponsored Products" not in wb.sheetnames:
        return
    ws = wb["Sponsored Products"]
    blocks = {"exact": 7, "phrase": 14, "broad": 21}
    kws = (keywords or [])[:6]
    for _match, start in blocks.items():
        for i, kw in enumerate(kws):
            ws.cell(row=start + i, column=4, value=kw)  # col D = Keyword


def _rescale_band_labels(ws) -> None:
    """Rewrite the template's static 'Score Levels' ranges to the /100 bands so
    the sheet's legend matches the rescaled scoring."""
    targets = [f"{lo}-{hi}" for lo, hi, _ in BANDS]
    replace = {}
    for old in (("0-75", "76-116", "117-159"), ("0-86", "87-134", "135-184")):
        replace.update({o: n for o, n in zip(old, targets)})
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in replace:
                c.value = replace[c.value.strip()]


# One-line "how it's scored" per rubric element, for the Instructions tab
# (documentation only — does not affect any scoring).
_ELEMENT_HOWTO = {
    "price": "Price positioning vs the other competitors — a competitively-priced listing scores higher (harder to beat).",
    "sponsored_products": "How aggressively the listing runs Sponsored Product ads across the launch keywords.",
    "sponsored_brands": "Presence of Sponsored Brand (headline) ad placements on the keywords.",
    "reviews": "Number of product reviews — more reviews = stronger social proof = harder to beat.",
    "rating": "Average star rating (higher rating scores higher).",
    "age": "Age of the listing — long-established listings score higher.",
    "product_images": "Main-gallery image quality & count: professional, lifestyle, infographics.",
    "product_title": "Title quality — relevant, clear, well-structured (not keyword-stuffed).",
    "bullets_description": "Bullet/description quality — benefits, keywords, addresses buyer pain points.",
    "bestseller": "Best Seller badge / BSR strength (badge held longer = higher).",
    "seller": "Amazon vs 3rd-party seller and seller feedback volume (established seller = higher).",
    "unique_design": "Design uniqueness — visibly differentiated (2) vs generic/saturated (0). Patents ignored.",
    "fba_fbm": "Fulfilment: FBA scores 2, FBM scores 0.",
    "mkt_images": "Overall brand/marketing imagery quality from the gallery + any A+ visible.",
    "mkt_title": "Marketing strength of the title (same listing, brand-marketing lens).",
    "mkt_bullets": "Marketing strength of the description/bullets.",
    "pricing_strategy": "Pricing strategy vs the market (positioning, discounts). -1/blank if not judgeable.",
    "review_velocity": "Rate of new reviews per month — fast velocity = strong momentum.",
    "variations": "Number of product variations offered (more variety = stronger).",
    "sales_strength": "Monthly sales percentile within the category pool.",
    "enhanced_content": "Enhanced media: A+ content and/or product video, else listing image count.",
}


def _write_instructions_sheet(wb) -> None:
    """Insert a first 'Instructions' tab explaining how each competitor-analysis
    element is scored and how the bands are read. Documentation only — reads the
    ELEMENTS/BANDS constants and writes a new sheet; touches no scoring."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    if "Instructions" in wb.sheetnames:
        del wb["Instructions"]
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 78

    head = PatternFill("solid", fgColor="12324C")
    band_fill = PatternFill("solid", fgColor="F4F7F9")
    thin = Side(style="thin", color="D8E1E8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wrap = Alignment(vertical="center", wrap_text=True)

    ws["A1"] = "Competitor Analysis — How Scoring Works"
    ws["A1"].font = Font(bold=True, size=16, color="12324C")
    ws.merge_cells("A1:D1")
    ws["A2"] = ("Each competitor is scored on 21 elements grouped in 4 sections. Scores are rescaled to a "
                "100-point total. A HIGHER score means a STRONGER competitor — i.e. HARDER for a new seller "
                "to beat. Any value that cannot be verified from a source is left blank / marked "
                "\"Unknown — need user confirmation\" (Rule 0: no guessing).")
    ws["A2"].alignment = wrap
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 46

    r = 4
    for col, title in zip("ABCD", ["Section", "Element", "Max pts", "How it is scored"]):
        c = ws[f"{col}{r}"]
        c.value = title
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = head
        c.border = border
        c.alignment = Alignment(vertical="center")
    r += 1
    last_section = None
    for key, label, mx, section in ELEMENTS:
        ws[f"A{r}"] = section if section != last_section else ""
        if section != last_section:
            ws[f"A{r}"].font = Font(bold=True, color="8A5B00")
        last_section = section
        ws[f"B{r}"] = label
        ws[f"C{r}"] = mx
        ws[f"D{r}"] = _ELEMENT_HOWTO.get(key, "")
        for col in "ABCD":
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].alignment = Alignment(vertical="center", wrap_text=True)
        r += 1

    ws[f"A{r}"] = "Raw max = 184 pts, rescaled proportionally to a 100-pt total (col K in each block)."
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"].font = Font(italic=True, size=9, color="5C7284")
    r += 2

    ws[f"A{r}"] = "Score bands (total /100)"
    ws[f"A{r}"].font = Font(bold=True, size=12, color="12324C")
    ws.merge_cells(f"A{r}:D{r}")
    r += 1
    for lo, hi, desc in BANDS:
        ws[f"A{r}"] = f"{lo}–{hi}"
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = desc
        ws.merge_cells(f"B{r}:D{r}")
        for col in "ABCD":
            ws[f"{col}{r}"].fill = band_fill
            ws[f"{col}{r}"].border = border
            ws[f"{col}{r}"].alignment = wrap
        r += 1

    r += 1
    ws[f"A{r}"] = ("Weaknesses / Solutions / Strengths come from the competitor's customer reviews "
                   "(1–2★ complaints and 4–5★ praise). With no review data they show "
                   "\"Unknown — need user confirmation\".")
    ws.merge_cells(f"A{r}:D{r}")
    ws[f"A{r}"].alignment = wrap
    ws.row_dimensions[r].height = 32


def write_competitor_workbook(
    template_path: str,
    top: List[Dict],
    scored_by_asin: Dict[str, Dict],
    reviews_by_asin: Dict[str, object],
    keywords: List[str],
    target_title: Optional[str] = None,
) -> str:
    """
    Fill the Competitor Analysis template for the top-N competitors and save a
    new timestamped workbook in output/. *top* is an ordered list of merged
    competitor record dicts (block 1 = top[0], etc.). Returns the saved path.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Competitor Analysis"]
    starts = _find_block_starts(ws)
    if not starts:
        raise ValueError("Could not locate competitor blocks in the template.")
    starts.append(ws.max_row + 1)  # sentinel end

    all_misses: List[str] = []
    for i, comp in enumerate(top):
        if i >= len(starts) - 1:
            break
        start, end = starts[i], starts[i + 1]
        scored = scored_by_asin.get(comp["asin"])
        review = reviews_by_asin.get(comp["asin"])
        if scored:
            all_misses += _fill_block(ws, start, end, comp, scored, review)

    _fill_pricing_analysis(wb, top, target_title)
    _fill_sponsored_products(wb, keywords)
    _rescale_band_labels(ws)
    _write_instructions_sheet(wb)  # first tab: how scoring works (docs only)

    out_dir = Path("output")
    out_dir.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = out_dir / f"Phase3_CompetitorAnalysis_{ts}.xlsx"
    wb.save(out_path)

    if all_misses:
        print("  NOTE (template rows not found — add them to fill):")
        for m in all_misses:
            print(f"    - {m}")
    return str(out_path)


# ==============================================================================
# PDF deliverable (reportlab) — a styled, downloadable competitor report
# ==============================================================================
def _pdf_fmt(value, kind="text"):
    """Format a record value for the PDF (currency/int/yesno/text)."""
    if value is None:
        return "—"
    if isinstance(value, float) and np.isnan(value):
        return "—"
    try:
        if kind == "money":
            return f"£{float(value):,.2f}"
        if kind == "int":
            return f"{int(round(float(value))):,}"
        if kind == "yesno":
            return "Yes" if value else "No"
        if kind == "rating":
            return f"{float(value):.1f}"
    except (ValueError, TypeError):
        return "—"
    s = str(value).strip()
    return s if s and s.lower() not in ("nan", "none") else "—"


def write_competitor_pdf(records, scored_by_asin, reviews_by_asin, kw_list,
                         target_title=None, out_dir="output"):
    """Write a styled Competitor Analysis PDF and return its path. Mirrors the
    Excel deliverable's data: per-competitor 184-point rubric + review insights."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                    TableStyle, HRFlowable)

    INDIGO = colors.HexColor("#4f46e5")
    INK = colors.HexColor("#1e293b")
    MUTED = colors.HexColor("#64748b")
    LINE = colors.HexColor("#e2e8f0")
    SOFT = colors.HexColor("#eef2ff")

    ss = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=ss["Title"], textColor=INK, fontSize=20, spaceAfter=2)
    sub = ParagraphStyle("sub", parent=ss["Normal"], textColor=MUTED, fontSize=9, spaceAfter=2)
    h2 = ParagraphStyle("h2", parent=ss["Heading2"], textColor=INDIGO, fontSize=13, spaceBefore=10, spaceAfter=4)
    sec = ParagraphStyle("sec", parent=ss["Normal"], textColor=MUTED, fontSize=8, spaceBefore=6, spaceAfter=2)
    body = ParagraphStyle("body", parent=ss["Normal"], textColor=INK, fontSize=9, leading=12)
    cell = ParagraphStyle("cell", parent=body, fontSize=8.5, leading=11)

    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"Competitor_Analysis_{date.today().isoformat()}.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=A4, title="Competitor Analysis",
                            leftMargin=16 * mm, rightMargin=16 * mm,
                            topMargin=16 * mm, bottomMargin=14 * mm)

    flow = [Paragraph("Competitor Analysis", h1)]
    if target_title:
        flow.append(Paragraph(str(target_title), sub))
    meta = f"Generated {date.today():%d %b %Y}"
    if kw_list:
        meta += "  ·  Launch keywords: " + ", ".join(str(k) for k in kw_list)
    flow += [Paragraph(meta, sub), Spacer(1, 4),
             HRFlowable(width="100%", thickness=1, color=LINE, spaceAfter=8)]

    # Summary table across all competitors.
    summary = [["#", "ASIN", "Brand", "Score", "Verdict"]]
    for i, rec in enumerate(records, 1):
        a = rec.get("asin")
        s = scored_by_asin.get(a, {})
        summary.append([str(i), str(a or "—"), Paragraph(str(rec.get("brand") or "—"), cell),
                        f"{s.get('total', '—')}/{s.get('max', MAX_TOTAL)}",
                        Paragraph(str(s.get("band") or "—"), cell)])
    st = Table(summary, colWidths=[8 * mm, 26 * mm, 36 * mm, 20 * mm, 88 * mm], repeatRows=1)
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), INDIGO), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, LINE), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    flow += [Paragraph("Summary", h2), st]

    # Per-competitor detail.
    for rec in records:
        a = rec.get("asin")
        s = scored_by_asin.get(a, {})
        flow.append(Paragraph(f"{rec.get('brand') or a}  —  {s.get('total', '—')}/{s.get('max', MAX_TOTAL)}", h2))
        flow.append(Paragraph(f"{a} · {s.get('band') or ''}", sub))

        facts = [
            ["Price", _pdf_fmt(rec.get("price"), "money"), "Sales/mo", _pdf_fmt(rec.get("monthly_sales"), "int")],
            ["Rating", _pdf_fmt(rec.get("rating"), "rating"), "Reviews", _pdf_fmt(rec.get("review_count"), "int")],
            ["BSR", _pdf_fmt(rec.get("bsr"), "int"), "Fulfillment", _pdf_fmt(rec.get("fulfillment"))],
            ["A+ content", _pdf_fmt(rec.get("a_plus"), "yesno"), "Video", _pdf_fmt(rec.get("has_video"), "yesno")],
        ]
        ft = Table(facts, colWidths=[26 * mm, 40 * mm, 26 * mm, 40 * mm])
        ft.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5), ("TEXTCOLOR", (0, 0), (0, -1), MUTED),
            ("TEXTCOLOR", (2, 0), (2, -1), MUTED), ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
            ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"), ("BACKGROUND", (0, 0), (-1, -1), SOFT),
            ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("BOX", (0, 0), (-1, -1), 0.5, LINE), ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.white),
        ]))
        flow += [Spacer(1, 4), ft]

        # Rubric grouped by section.
        elements = s.get("elements", {})
        rubric = [["Element", "Score"]]
        spans = []
        last_section = None
        for key, label, mx, section in ELEMENTS:
            if key not in elements:
                continue
            if section != last_section:
                spans.append(len(rubric))
                rubric.append([section, ""])
                last_section = section
            sc = elements[key].get("score")
            emax = elements[key].get("max", mx)  # rescaled /100 weight
            rubric.append([Paragraph(label, cell), f"{'—' if sc is None else sc}/{emax}"])
        rt = Table(rubric, colWidths=[112 * mm, 26 * mm], repeatRows=1)
        style = [
            ("BACKGROUND", (0, 0), (-1, 0), INK), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"), ("LINEBELOW", (0, 1), (-1, -1), 0.4, LINE),
            ("TOPPADDING", (0, 0), (-1, -1), 2.5), ("BOTTOMPADDING", (0, 0), (-1, -1), 2.5),
        ]
        for r in spans:
            style += [("SPAN", (0, r), (-1, r)), ("BACKGROUND", (0, r), (-1, r), SOFT),
                      ("FONTNAME", (0, r), (0, r), "Helvetica-Bold"), ("TEXTCOLOR", (0, r), (0, r), INDIGO)]
        rt.setStyle(TableStyle(style))
        flow += [Spacer(1, 5), rt]

        # Review insights.
        ra = reviews_by_asin.get(a)
        strengths = list(getattr(ra, "strengths", []) or [])
        weaknesses = list(getattr(ra, "weaknesses", []) or [])
        solutions = list(getattr(ra, "solutions", []) or [])
        if strengths or weaknesses:
            flow.append(Paragraph("Review insights", sec))
            if strengths:
                flow.append(Paragraph("<b>Strengths:</b> " + "; ".join(strengths), body))
            for i, w in enumerate(weaknesses):
                fix = f"  <font color='#64748b'>→ fix: {solutions[i]}</font>" if i < len(solutions) else ""
                flow.append(Paragraph(f"<b>Weakness:</b> {w}{fix}", body))
        flow.append(Spacer(1, 8))

    doc.build(flow)
    return str(path)


# ──────────────────────────────────────────────────────────────────────────────
# High-level entry point (non-interactive; returns data, writes the workbook)
# ──────────────────────────────────────────────────────────────────────────────

DEFAULT_TEMPLATE = "test_file/competitor_analysis/Compatitor Analysis.xlsx"


def _lookup_title(asin, js_df, per_kw, blackbox_df=None):
    """Find a title + category for the target ASIN across the provided files."""
    for df in [blackbox_df, js_df, *per_kw.values()]:
        if df is None or "asin" not in getattr(df, "columns", []):
            continue
        m = df[df["asin"].astype(str).str.upper() == asin]
        if not m.empty:
            row = m.iloc[0]
            return {
                "title": str(row.get("title") or "").strip(),
                "category": str(row.get("category") or "").strip(),
            }
    return {}


def _build_per_kw(xray_files):
    """Accept a dict {keyword: path}, a list of paths, or a folder / comma-list
    string, and return {keyword: DataFrame} of Xray Products exports."""
    per_kw = {}
    if isinstance(xray_files, dict):
        for kw, fp in xray_files.items():
            df, _ = read_xray_products(str(fp))
            per_kw[str(kw)] = df
        return per_kw
    if isinstance(xray_files, (list, tuple)):
        paths = [Path(str(p)) for p in xray_files]
    elif isinstance(xray_files, str):
        p = Path(xray_files)
        if p.is_dir():
            paths = sorted(p.glob("Xray_products_*.xlsx")) or sorted(p.glob("*.xlsx"))
        else:
            paths = [Path(x.strip().strip('"').strip("'")) for x in xray_files.split(",") if x.strip()]
    else:
        return per_kw
    for fp in paths:
        kw = fp.stem.replace("Xray_products_Keyword_", "").replace("_", " ").strip()
        df, _ = read_xray_products(str(fp))
        per_kw[kw] = df
    return per_kw


def run_phase3(target_asin, js_file=None, xray_files=None, *, locked_keywords=None,
               js_autogen=None, blackbox_df=None, template=DEFAULT_TEMPLATE,
               deep=True, deep_limit=None, reviews_by_asin=None, write_workbook=True):
    """Run Phase 3 (competitor identification + analysis) for one target ASIN and
    return data. Inputs are passed in (no prompts): the auto-generated/uploaded
    Jungle Scout sheet (`js_file` path, or join an in-flight `js_autogen` job from
    Phase 2) and one H10 Xray *Products* export per launch keyword (`xray_files`:
    dict {keyword: path}, list of paths, or a folder/comma-list string). The
    target's title/category are read from those files. Claude builds the profile,
    selects the top-3 same-type competitors, Keepa deep-pulls them, reviews come
    from the FlyBy API, and the 184-point rubric is scored. The Competitor
    Analysis workbook is written unless `write_workbook=False`. Errors are
    returned in the `error` key (never raised)."""
    log: list = []
    out = {
        "target_asin": (target_asin or "").strip().upper(),
        "profile": None, "top": [], "same_type_count": None, "scored": [],
        "records": {}, "workbook_path": None, "pdf_path": None, "log": log, "error": None,
    }
    asin = out["target_asin"]
    if not asin:
        out["error"] = "No target ASIN supplied."
        return out

    try:
        per_kw = _build_per_kw(xray_files)
    except Exception as e:  # noqa: BLE001
        out["error"] = f"Could not read the Xray product exports: {e}"
        return out
    if not per_kw:
        out["error"] = "No Xray product exports provided (one per launch keyword)."
        return out

    # Jungle Scout sheet: join an in-flight Phase 2 job, else read js_file.
    js_df = None
    try:
        if js_autogen:
            thread, holder = js_autogen["job"]
            thread.join()
            if "res" in holder:
                log.append(f"JS sheet ready: {holder['res']['out_path']} ({holder['res'].get('count')} products)")
                js_df, _ = read_js_products(holder["res"]["out_path"])
            else:
                log.append(f"JS-sheet generation failed: {holder.get('err')}")
        if js_df is None and js_file:
            js_df, _ = read_js_products(str(js_file))
    except Exception as e:  # noqa: BLE001
        out["error"] = f"Could not read the Jungle Scout sheet: {e}"
        return out
    if js_df is None:
        out["error"] = "No Jungle Scout data — provide js_file or a js_autogen job."
        return out

    try:
        client = load_client()
    except Phase3ApiError as e:
        out["error"] = str(e)
        return out

    info = _lookup_title(asin, js_df, per_kw, blackbox_df)
    if not info.get("title"):
        out["error"] = f"Could not find a title for {asin} in the provided files."
        return out
    try:
        profile = derive_product_profile(client, title=info["title"], category=info.get("category", ""))
    except Phase3ApiError as e:
        out["error"] = str(e)
        return out
    out["profile"] = profile.model_dump()

    # Step 8 — pool -> same-type -> top 3.
    try:
        agg = aggregate_sponsored(per_kw)
        pool = build_asin_pool(js_df, list(per_kw.values()))
        records = build_candidate_records(js_df, per_kw, agg, pool)
        top, same_type, report = select_top_competitors(client, profile, records, n=3)
    except (Phase3ApiError, ValueError) as e:
        out["error"] = str(e)
        return out
    log.extend(report)
    if top.empty:
        out["error"] = "No same-product-type competitors found."
        return out

    top_asins = list(top["asin"])
    records_by_asin = {r["asin"]: dict(r) for _, r in top.iterrows()}

    # Step 10A — Keepa deep pull (final 3 only).
    try:
        api = load_keepa()
        products = query_products(api, top_asins, deep=deep, limit=deep_limit)
        log.append(f"Keepa returned {len(products)} of {len(top_asins)} (tokens left: {api.tokens_left})")
        for a in top_asins:
            if a in products:
                f = extract_keepa_fields(products[a])
                if f.get("buybox_seller_id"):
                    f["seller_feedback"] = get_seller_feedback(api, f["buybox_seller_id"])
                records_by_asin[a].update({k: v for k, v in f.items()
                                           if v is not None or k in ("a_plus", "has_video", "seller_feedback")})
    except Phase3KeepaError as e:
        log.append(f"Keepa unavailable: {e} (variations/A+/seller-feedback Unknown).")

    # Step 10C — Claude listing scoring + reviews, then the rubric.
    pool_stats = pool_statistics(same_type)
    reviews_in = reviews_by_asin or {}
    scored: list = []
    for a in top_asins:
        rec = records_by_asin[a]
        try:
            listing = score_listing(client, {
                "title": rec.get("title"), "bullets": rec.get("bullets"),
                "price": rec.get("price"), "a_plus": rec.get("a_plus"),
                "has_video": rec.get("has_video"), "images_count": rec.get("images_count"),
                "market_price_context": f"prices: {sorted([p for p in top['price'] if pd.notna(p)])}",
            }, image_urls=rec.get("image_urls"))
        except Phase3ApiError as e:
            log.append(f"Listing scoring failed for {a}: {e}")
            listing = None
        rev_text = reviews_in.get(a, "")
        if not rev_text:
            rev = fetch_reviews(a)
            if rev.get("error"):
                log.append(f"Reviews {a} failed: {rev['error']}")
            rec["review_counts"] = {
                "total_reviews": rev.get("total_ratings"),
                "distribution": rev.get("distribution"),
                "weaknesses_count": rev.get("weaknesses_count"),
            }
            rev_text = reviews_text(rev)
        rec["review_analysis"] = analyze_reviews(client, rec.get("brand") or a, rev_text)
        scored.append(score_competitor(rec, [dict(r) for _, r in top.iterrows()], pool_stats, listing))

    # Step 10E — write the Competitor Analysis deliverables (Excel + PDF).
    if write_workbook:
        kw_list = [getattr(k, "keyword", k) for k in (locked_keywords or [])]
        review_analyses = {a: records_by_asin[a].get("review_analysis") for a in top_asins}
        comp_records = [records_by_asin[a] for a in top_asins]
        scored_map = {s["asin"]: s for s in scored}
        try:
            out["workbook_path"] = write_competitor_workbook(
                template, comp_records, scored_map, review_analyses, kw_list,
                target_title=info.get("title"),
            )
        except Exception as e:  # noqa: BLE001
            log.append(f"Could not write workbook: {e}")
        try:
            out["pdf_path"] = write_competitor_pdf(
                comp_records, scored_map, review_analyses, kw_list,
                target_title=info.get("title"),
            )
        except Exception as e:  # noqa: BLE001
            log.append(f"Could not write PDF: {e}")

    out["top"] = [dict(r) for _, r in top.iterrows()]
    out["same_type_count"] = int(len(same_type)) if same_type is not None else None
    out["scored"] = scored
    out["records"] = {a: {k: v for k, v in records_by_asin[a].items() if k != "review_analysis"}
                      for a in top_asins}
    return out
